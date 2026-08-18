#!/usr/bin/env python3
"""ตรวจสอบอิสระว่าไฟล์ตัวอย่างที่สร้างออกมา มีกับดักอยู่จริงตามที่ประกาศไว้"""
import json, sys, os, re
import openpyxl

d = sys.argv[1]
meta = json.load(open("samples_meta.json", encoding="utf-8"))
log = json.load(open("samples_trap_log.json", encoding="utf-8"))
fail = 0

def num_like(v):
    return isinstance(v, str) and re.fullmatch(r"-?[\d,]+(\.\d+)?", v.strip()) is not None

for m in meta:
    path = os.path.join(d, m["fname"])
    wb = openpyxl.load_workbook(path, data_only=True)
    assert "ข้อมูลดิบ" in wb.sheetnames, m["fname"]
    ws = wb["ข้อมูลดิบ"]
    hdr = [c.value for c in ws[1]]
    data = list(ws.iter_rows(min_row=2, values_only=True))

    n_declared = None
    # อ่านจำนวนกับดักที่ประกาศไว้ในชีตอ่านก่อนเริ่ม
    intro = wb["อ่านก่อนเริ่ม"]
    for row in intro.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and "ไฟล์นี้มีปัญหาข้อมูลซ่อนอยู่" in v:
                n_declared = int(re.search(r"(\d+)", v).group(1))
    n_actual = len(log[m["code"]])

    # ตรวจของจริงในไฟล์
    checks = {}
    flat = [str(c) for row in data for c in row if c is not None]
    checks["มีข้อความที่หน้าตาเป็นตัวเลข"] = sum(1 for row in data for c in row if num_like(str(c)) and isinstance(c, str))
    checks["มีค่า -999"] = sum(1 for row in data for c in row if c == -999)
    checks["มีเซลล์ที่มีช่องว่างหัวหรือท้าย"] = sum(
        1 for row in data for c in row if isinstance(c, str) and c != c.strip() and c.strip() != "")
    checks["มีค่าว่าง"] = sum(1 for row in data for c in row if c is None or c == "")
    checks["มีวันที่แบบ พ.ศ."] = sum(1 for row in data for c in row
                                     if isinstance(c, str) and re.fullmatch(r"\d{2}/\d{2}/25\d{2}", c.strip() or ""))
    seen, dups = set(), 0
    for row in data:
        t = tuple(row)
        if t in seen:
            dups += 1
        seen.add(t)
    checks["แถวซ้ำ"] = dups

    ok = (n_declared == n_actual)
    status = "OK " if ok else "MISMATCH"
    if not ok:
        fail += 1
    print("{} {:3} {:38} แถว {:>4} | ประกาศ {} ประเภท / บันทึก {} ประเภท".format(
        status, m["code"], m["fname"][:38], len(data), n_declared, n_actual))
    print("      ตรวจพบจริงในไฟล์:", ", ".join("{}={}".format(k, v) for k, v in checks.items() if v))

    # ตรวจสิ่งที่ต้องไม่มี
    if len(hdr) != len(set(hdr)):
        print("      !! หัวตารางซ้ำกัน"); fail += 1
    if any(c is None for c in hdr):
        print("      !! หัวตารางมีช่องว่าง"); fail += 1

# ตรวจกับดักเฉพาะสาขา
print("\n--- ตรวจกับดักประจำสาขา ---")
wb = openpyxl.load_workbook(os.path.join(d, "AC_สมุดรายวันทั่วไป.xlsx"), data_only=True)
ws = wb["ข้อมูลดิบ"]
hdr = [c.value for c in ws[1]]
iv, idb, icr = hdr.index("เลขที่ใบสำคัญ"), hdr.index("เดบิต"), hdr.index("เครดิต")
bal = {}
def tonum(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s.startswith("(") and s.endswith(")"):
        return -float(s[1:-1])
    try:
        return float(s)
    except ValueError:
        return 0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    v = str(row[iv]).strip()
    bal[v] = bal.get(v, 0) + tonum(row[idb]) - tonum(row[icr])
unbal = [k for k, x in bal.items() if abs(x) > 0.005]
print("AC  ใบสำคัญที่ไม่ดุล: {} ใบ -> {}".format(len(unbal), sorted(unbal)[:10]))

wb = openpyxl.load_workbook(os.path.join(d, "GS_ข้อมูลเจาะสำรวจน้ำบาดาล.xlsx"), data_only=True)
ws = wb["ข้อมูลดิบ"]
hdr = [c.value for c in ws[1]]
it, ib = hdr.index("ความลึกช่วงบน (ม.)"), hdr.index("ความลึกช่วงล่าง (ม.)")
bad = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
          if isinstance(r[it], (int, float)) and isinstance(r[ib], (int, float)) and r[it] > r[ib])
print("GS  แถวที่ความลึกบน > ล่าง: {} แถว".format(bad))

wb = openpyxl.load_workbook(os.path.join(d, "FT_บันทึกควบคุมคุณภาพสายการผลิต.xlsx"), data_only=True)
ws = wb["ข้อมูลดิบ"]
hdr = [c.value for c in ws[1]]
im = hdr.index("จุลินทรีย์ทั้งหมด (CFU/g)")
txt = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if isinstance(r[im], str))
print("FT  ผลจุลินทรีย์ที่เป็นข้อความ: {} แถว".format(txt))

wb = openpyxl.load_workbook(os.path.join(d, "ED_สถานีตรวจวัดสิ่งแวดล้อม.xlsx"), data_only=True)
ws = wb["ข้อมูลดิบ"]
hdr = [c.value for c in ws[1]]
ip = hdr.index("PM2.5 (µg/m³)")
neg = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[ip] == -999)
vals = [r[ip] for r in ws.iter_rows(min_row=2, values_only=True) if isinstance(r[ip], (int, float))]
raw_avg = sum(vals) / len(vals)
clean = [v for v in vals if v != -999]
print("ED  PM2.5: มี -999 {} แถว | ค่าเฉลี่ยถ้าไม่ล้าง = {:.1f} | ล้างแล้ว = {:.1f}".format(
    neg, raw_avg, sum(clean) / len(clean)))

wb = openpyxl.load_workbook(os.path.join(d, "BA_ยอดขายร้านกาแฟ.xlsx"), data_only=True)
ws = wb["ข้อมูลดิบ"]
hdr = [c.value for c in ws[1]]
isc = hdr.index("สาขา")
branches = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    branches[str(r[isc])] = branches.get(str(r[isc]), 0) + 1
print("BA  ชื่อสาขาที่ปรากฏ: {} แบบ -> {}".format(len(branches), list(branches.keys())))

print("\nสรุป:", "ผ่านทั้งหมด" if fail == 0 else "พบปัญหา {} จุด".format(fail))
sys.exit(1 if fail else 0)
