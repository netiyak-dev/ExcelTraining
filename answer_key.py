#!/usr/bin/env python3
"""
สร้างเฉลยจาก "ไฟล์ที่ส่งมอบจริง" ไม่ใช่จากบันทึกตอนสร้าง
เพื่อให้ตัวเลขในคู่มืออาจารย์ตรงกับสิ่งที่ผู้เรียนจะเห็นเสมอ
"""
import json, os, re, sys
from collections import Counter
import openpyxl

D = sys.argv[1]
meta = json.load(open("samples_meta.json", encoding="utf-8"))
log = json.load(open("samples_trap_log.json", encoding="utf-8"))
KEY = {}


def load(fn):
    wb = openpyxl.load_workbook(os.path.join(D, fn), data_only=True)
    ws = wb["ข้อมูลดิบ"]
    hdr = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return wb, hdr, rows


def col(hdr, rows, name):
    i = hdr.index(name)
    return [r[i] for r in rows]


def tonum(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s.startswith("(") and s.endswith(")"):
        try:
            return -float(s[1:-1])
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def dup_count(rows):
    seen, n = set(), 0
    for r in rows:
        if tuple(r) in seen:
            n += 1
        seen.add(tuple(r))
    return n


def generic(code, fn):
    wb, hdr, rows = load(fn)
    txt_num = sum(1 for r in rows for c in r
                  if isinstance(c, str) and re.fullmatch(r"-?[\d,]+(\.\d+)?", c.strip() or "x"))
    pad = sum(1 for r in rows for c in r if isinstance(c, str) and c and c != c.strip())
    blank = sum(1 for r in rows for c in r if c is None or c == "")
    be = sum(1 for r in rows for c in r
             if isinstance(c, str) and re.fullmatch(r"\d{2}/\d{2}/25\d{2}", c.strip() or "x"))
    return {
        "จำนวนแถวข้อมูล": len(rows),
        "แถวซ้ำทั้งแถว": dup_count(rows),
        "เซลล์ที่เป็นข้อความแต่หน้าตาเป็นตัวเลข": txt_num,
        "เซลล์ที่มีช่องว่างนำหน้าหรือตามหลัง": pad,
        "เซลล์ว่าง": blank,
        "วันที่เป็น พ.ศ. แบบข้อความ": be,
    }, hdr, rows, wb


# ---------------- AG ----------------
g, hdr, rows, wb = generic("AG", "AG_แปลงทดลองข้าวโพด.xlsx")
y = col(hdr, rows, "ผลผลิต (กก.)")
g["ค่า -999 ในคอลัมน์ผลผลิต"] = sum(1 for v in y if v == -999)
vals = [tonum(v) for v in y]
vals = [v for v in vals if v is not None]
clean = [v for v in vals if v != -999]
g["ผลของการไม่ล้างข้อมูล"] = "ค่าเฉลี่ยผลผลิตจากข้อมูลดิบ {:,.0f} กก. เทียบกับ {:,.0f} กก. เมื่อตัด -999 ออก".format(
    sum(vals) / len(vals), sum(clean) / len(clean))
mo = col(hdr, rows, "ความชื้นเมล็ด (%)")
g["ค่าความชื้น 0 ซึ่งเป็นไปไม่ได้"] = sum(1 for v in mo if v == 0)
ref = wb["ตารางพันธุ์"]
known = set(str(r[0]).strip() for r in ref.iter_rows(min_row=2, values_only=True) if r[0])
used = set(str(v).strip() for v in col(hdr, rows, "รหัสพันธุ์"))
g["รหัสพันธุ์ที่ไม่มีในตารางอ้างอิง"] = sorted(used - known)
g["ชื่อตำบลที่ปรากฏทั้งหมด"] = len(set(str(v) for v in col(hdr, rows, "ตำบล")))
g["ชื่อตำบลหลังตัดช่องว่างและจุด"] = len(set(str(v).strip().rstrip(".") for v in col(hdr, rows, "ตำบล")))
KEY["AG"] = g

# ---------------- AC ----------------
g, hdr, rows, wb = generic("AC", "AC_สมุดรายวันทั่วไป.xlsx")
iv, idb, icr = hdr.index("เลขที่ใบสำคัญ"), hdr.index("เดบิต"), hdr.index("เครดิต")
def imbalance(rs):
    bal = {}
    for r in rs:
        v = str(r[iv]).strip()
        bal[v] = bal.get(v, 0) + (tonum(r[idb]) or 0) - (tonum(r[icr]) or 0)
    return sorted(k for k, x in bal.items() if abs(x) > 0.005)
seen, dedup = set(), []
for r in rows:
    if tuple(r) not in seen:
        dedup.append(r)
    seen.add(tuple(r))
before, after = imbalance(rows), imbalance(dedup)
g["ใบสำคัญไม่ดุลถ้ายังไม่ลบแถวซ้ำ"] = len(before)
g["ใบสำคัญไม่ดุลจริงหลังลบแถวซ้ำ"] = len(after)
g["เลขที่ใบสำคัญที่ไม่ดุลจริง"] = after
g["บทเรียนจากลำดับการทำงาน"] = ("ถ้าตรวจดุลก่อนลบแถวซ้ำ จะพบใบสำคัญผิดปกติ {} ใบ "
                                  "แต่ที่ผิดจริงมีเพียง {} ใบ ลำดับการทำความสะอาดข้อมูลจึงเปลี่ยนข้อสรุปได้".format(
                                      len(before), len(after)))
g["จำนวนเงินที่อยู่ในวงเล็บ"] = sum(1 for r in rows for c in r
                                     if isinstance(c, str) and c.strip().startswith("(") and c.strip().endswith(")"))
ref = wb["ผังบัญชี"]
known = set(str(r[0]).strip() for r in ref.iter_rows(min_row=2, values_only=True) if r[0])
used = set(str(v).strip() for v in col(hdr, rows, "รหัสบัญชี"))
g["รหัสบัญชีที่ไม่มีในผังบัญชี"] = sorted(used - known)
KEY["AC"] = g

# ---------------- BA ----------------
g, hdr, rows, wb = generic("BA", "BA_ยอดขายร้านกาแฟ.xlsx")
br = Counter(str(v) for v in col(hdr, rows, "สาขา"))
g["ชื่อสาขาที่ปรากฏ"] = dict(br)
g["จำนวนสาขาจริง"] = len(set(s.strip().replace("สาขา", "").replace(" ", "") for s in br))
disc = col(hdr, rows, "ส่วนลด (บาท)")
g["ส่วนลดที่กรอกเป็นเปอร์เซ็นต์"] = sum(1 for v in disc if isinstance(v, str) and "%" in v)
qty = col(hdr, rows, "จำนวน")
g["รายการที่จำนวนติดลบ (คืนสินค้า ไม่ใช่ข้อผิดพลาด)"] = sum(1 for v in qty if isinstance(v, (int, float)) and v < 0)
ref = wb["ตารางสินค้า"]
known = set(str(r[0]).strip() for r in ref.iter_rows(min_row=2, values_only=True) if r[0])
used = set(str(v).strip() for v in col(hdr, rows, "รหัสสินค้า"))
g["รหัสสินค้าที่ไม่มีในตารางสินค้า"] = sorted(used - known)
KEY["BA"] = g

# ---------------- CB ----------------
g, hdr, rows, wb = generic("CB", "CB_สำรวจกล้องดักถ่าย.xlsx")
q = col(hdr, rows, "จำนวนตัว")
g["จำนวนที่กรอกเป็นคำบรรยาย"] = sum(1 for v in q if isinstance(v, str))
g["ตัวอย่างคำที่พบ"] = sorted(set(str(v) for v in q if isinstance(v, str)))
t = col(hdr, rows, "อุณหภูมิ (°C)")
g["อุณหภูมิ 99.9 ซึ่งเป็นค่าผิดปกติของเซนเซอร์"] = sum(1 for v in t if v == 99.9)
ref = wb["ตารางชนิด"]
known = set(str(r[0]).strip() for r in ref.iter_rows(min_row=2, values_only=True) if r[0])
used = set(str(v).strip() for v in col(hdr, rows, "รหัสชนิด"))
g["รหัสชนิดที่ไม่มีในตารางชนิด"] = sorted(used - known)
tm = col(hdr, rows, "เวลา")
g["เวลาที่ใช้จุดแทนทวิภาค"] = sum(1 for v in tm if isinstance(v, str) and "." in str(v))
g["ผลของแถวซ้ำ"] = "ถ้าไม่ลบแถวซ้ำ ความถี่การพบจะสูงเกินจริง {} ครั้ง จากทั้งหมด {} แถว".format(
    g["แถวซ้ำทั้งแถว"], g["จำนวนแถวข้อมูล"])
KEY["CB"] = g

# ---------------- ED ----------------
g, hdr, rows, wb = generic("ED", "ED_สถานีตรวจวัดสิ่งแวดล้อม.xlsx")
p = [tonum(v) for v in col(hdr, rows, "PM2.5 (µg/m³)")]
p = [v for v in p if v is not None]
g["ค่า -999 ใน PM2.5"] = sum(1 for v in p if v == -999)
g["ผลของการไม่ล้างข้อมูล"] = "ค่าเฉลี่ย PM2.5 จากข้อมูลดิบ = {:.1f} (ติดลบ) เทียบกับ {:.1f} เมื่อตัด -999 ออก".format(
    sum(p) / len(p), sum(v for v in p if v != -999) / len([v for v in p if v != -999]))
p10 = col(hdr, rows, "PM10 (µg/m³)")
g["ค่า 9999 ใน PM10 (เซนเซอร์ผิดพลาด)"] = sum(1 for v in p10 if v == 9999)
r_ = col(hdr, rows, "ปริมาณน้ำฝน (มม.)")
g["ค่า -999 ในปริมาณน้ำฝน"] = sum(1 for v in r_ if v == -999)
# ตรวจวันที่ขาดหายของอนุกรมเวลา
import datetime
def parse_d(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    m = re.fullmatch(r"(\d{2})/(\d{2})/(25\d{2})", s)
    if m:
        return datetime.date(int(m.group(3)) - 543, int(m.group(2)), int(m.group(1)))
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None
ds = [parse_d(v) for v in col(hdr, rows, "วันที่")]
ds = sorted(set(d for d in ds if d))
span = (ds[-1] - ds[0]).days + 1
g["ช่วงวันที่ของข้อมูล"] = "{} ถึง {} รวม {} วันตามปฏิทิน".format(ds[0], ds[-1], span)
g["จำนวนวันที่มีข้อมูลจริง"] = len(ds)
g["จำนวนวันที่ขาดหายไป"] = span - len(ds)
KEY["ED"] = g

# ---------------- FT ----------------
g, hdr, rows, wb = generic("FT", "FT_บันทึกควบคุมคุณภาพสายการผลิต.xlsx")
mb = col(hdr, rows, "จุลินทรีย์ทั้งหมด (CFU/g)")
g["ผลจุลินทรีย์ที่เป็นข้อความ"] = sum(1 for v in mb if isinstance(v, str))
g["รูปแบบข้อความที่พบ"] = sorted(set(str(v) for v in mb if isinstance(v, str)))
w = [tonum(v) for v in col(hdr, rows, "น้ำหนักสุทธิ (ก.)")]
g["แถวที่น้ำหนักน่าจะเป็นกิโลกรัม (ค่าน้อยกว่า 10)"] = sum(1 for v in w if v is not None and v < 10)
ref = wb["สเปกผลิตภัณฑ์"]
spec = {str(r[0]).strip(): r for r in ref.iter_rows(min_row=2, values_only=True) if r[0]}
ip, iph, iw = hdr.index("รหัสผลิตภัณฑ์"), hdr.index("pH"), hdr.index("น้ำหนักสุทธิ (ก.)")
out_ph = 0
for r in rows:
    code = str(r[ip]).strip()
    ph = tonum(r[iph])
    if code in spec and ph is not None and not (spec[code][2] <= ph <= spec[code][3]):
        out_ph += 1
g["ล็อตที่ pH หลุดสเปก (นับหลังแปลงข้อความเป็นตัวเลขแล้ว)"] = out_ph
known = set(spec)
used = set(str(v).strip() for v in col(hdr, rows, "รหัสผลิตภัณฑ์"))
g["รหัสผลิตภัณฑ์ที่ไม่มีในสเปก"] = sorted(used - known)
KEY["FT"] = g

# ---------------- GS ----------------
g, hdr, rows, wb = generic("GS", "GS_ข้อมูลเจาะสำรวจน้ำบาดาล.xlsx")
it, ib = hdr.index("ความลึกช่วงบน (ม.)"), hdr.index("ความลึกช่วงล่าง (ม.)")
bad = [i + 2 for i, r in enumerate(rows)
       if isinstance(r[it], (int, float)) and isinstance(r[ib], (int, float)) and r[it] > r[ib]]
g["แถวที่ความลึกบนมากกว่าล่าง"] = len(bad)
g["ตัวอย่างแถวที่ผิดตรรกะ"] = bad[:12]
tds = col(hdr, rows, "TDS (mg/L)")
g["ค่า TDS 99999 (ค่าผิดปกติ)"] = sum(1 for v in tds if v == 99999)
u = Counter(str(v).strip() for v in col(hdr, rows, "รหัสหน่วยหิน"))
g["รหัสหน่วยหินที่ปรากฏ"] = dict(u)
g["รหัสหน่วยหินจริงหลังปรับตัวพิมพ์"] = len(set(k.upper() for k in u))
e = col(hdr, rows, "พิกัด E")
g["พิกัดที่เก็บเป็นข้อความ"] = sum(1 for v in e if isinstance(v, str))
q = col(hdr, rows, "ปริมาณน้ำ (ลบ.ม./ชม.)")
g["หลุมที่ปริมาณน้ำเป็น 0 (ต้องตีความว่าเจาะไม่พบน้ำ ไม่ใช่ข้อมูลหาย)"] = sum(1 for v in q if v == 0)
KEY["GS"] = g

json.dump(KEY, open("answer_key.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
for k, v in KEY.items():
    print("\n===", k, "===")
    for kk, vv in v.items():
        s = str(vv)
        print("  {:55} {}".format(kk, s if len(s) < 90 else s[:90] + "..."))
