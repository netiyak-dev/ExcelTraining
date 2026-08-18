#!/usr/bin/env python3
"""สร้างไฟล์ตัวอย่าง .xlsx 7 สาขา พร้อมกับดักที่ตั้งใจซ่อนไว้ และบันทึกเฉลยเป็น JSON"""
import json, random, sys, os
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datasets as DS

random.seed(DS.SEED + 7)

FONT = "Tahoma"
NAVY = "1F3864"; HEAD = "D9E2F3"; ORANGE = "FBE5D6"; GREEN = "E2EFDA"; GREY = "F2F2F2"
OUT = sys.argv[1] if len(sys.argv) > 1 else "."

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

LOG = {}


def log(ds, trap, detail):
    LOG.setdefault(ds, {}).setdefault(trap, []).append(detail)


# ============================================================ กับดักมาตรฐาน
def trap_text_numbers(rows, col, n, ds):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = DS.as_text(rows[i][col])
        log(ds, "ตัวเลขถูกเก็บเป็นข้อความ", "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_padded(rows, col, n, ds):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = DS.pad(rows[i][col])
        log(ds, "มีช่องว่างนำหน้าหรือตามหลัง", "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_thai_date(rows, col, n, ds):
    import datetime
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        y, m, d = [int(x) for x in str(rows[i][col]).split("-")]
        rows[i][col] = DS.th_date(datetime.date(y, m, d))
        log(ds, "วันที่เป็น พ.ศ. แบบข้อความ", "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_blank(rows, col, n, ds):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = ""
        log(ds, "ค่าว่าง (missing value)", "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_sentinel(rows, col, n, ds, value=-999):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = value
        log(ds, "ค่าแทนการไม่มีข้อมูล ({})".format(value), "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_unit_in_cell(rows, col, n, ds, unit):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = "{} {}".format(rows[i][col], unit)
        log(ds, "หน่วยติดมากับตัวเลขในเซลล์เดียวกัน", "แถว {} คอลัมน์ {}".format(i + 2, col))


def trap_variant_spelling(rows, col, mapping, n, ds):
    cand = [i for i, r in enumerate(rows) if r[col] in mapping]
    idx = random.sample(cand, min(n, len(cand)))
    for i in idx:
        old = rows[i][col]
        rows[i][col] = mapping[old]
        log(ds, "สะกดชื่อไม่ตรงกัน", "แถว {} คอลัมน์ {}: {} → {}".format(i + 2, col, old, rows[i][col]))


def trap_duplicate_rows(rows, n, ds):
    idx = random.sample(range(len(rows)), n)
    for i in sorted(idx, reverse=True):
        rows.insert(i + 1, deepcopy(rows[i]))
        log(ds, "แถวซ้ำทั้งแถว", "แถวประมาณที่ {}".format(i + 2))


def trap_orphan_key(rows, col, n, ds, value):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = value
        log(ds, "รหัสที่ไม่มีในตารางอ้างอิง", "แถว {} คอลัมน์ {} = {}".format(i + 2, col, value))


def trap_outlier(rows, col, n, ds, value):
    idx = random.sample(range(len(rows)), n)
    for i in idx:
        rows[i][col] = value
        log(ds, "ค่าผิดปกติจากเครื่องมือหรือการบันทึก", "แถว {} คอลัมน์ {} = {}".format(i + 2, col, value))


# ============================================================ นิยามชุดข้อมูล
def build_all():
    specs = []

    # ---------------- AG ----------------
    rows, ref = DS.ag()
    trap_text_numbers(rows, "ผลผลิต (กก.)", 22, "AG")
    trap_unit_in_cell(rows, "พื้นที่ (ไร่)", 14, "AG", "ไร่")
    trap_thai_date(rows, "วันที่ปลูก", 26, "AG")
    trap_variant_spelling(rows, "ตำบล", {"ลุ่มสุ่ม": "ลุ่มสุ่ม ", "หนองโรง": " หนองโรง",
                                          "วังด้ง": "วังด้ง.", "แก่งเสี้ยน": "แก่งเสี้ยน "}, 18, "AG")
    trap_sentinel(rows, "ผลผลิต (กก.)", 8, "AG")
    trap_blank(rows, "ความชื้นเมล็ด (%)", 12, "AG")
    trap_outlier(rows, "ความชื้นเมล็ด (%)", 5, "AG", 0)
    trap_orphan_key(rows, "รหัสพันธุ์", 6, "AG", "CP-777")
    trap_duplicate_rows(rows, 9, "AG")
    specs.append(dict(
        code="AG", prog="วิทยาศาสตร์การเกษตร",
        fname="AG_แปลงทดลองข้าวโพด.xlsx",
        title="บันทึกแปลงทดลองข้าวโพดเลี้ยงสัตว์ ฤดูกาลผลิต 2568/69",
        context="ข้อมูลบันทึกแปลงทดลอง 280 แปลงใน 6 ตำบลของจังหวัดกาญจนบุรี รวบรวมโดยผู้บันทึก 4 คน "
                "ซึ่งใช้รูปแบบการกรอกต่างกัน จึงมีความไม่สม่ำเสมอแบบที่พบจริงในงานภาคสนาม",
        rows=rows, refs=[ref], traps=9,
        tasks=[
            ("Beginner",
             "ทำความสะอาดข้อมูลให้พร้อมวิเคราะห์ แล้วตอบว่าตำบลใดมีผลผลิตรวมสูงที่สุด",
             "ต้องแปลงผลผลิตที่เป็นข้อความให้เป็นตัวเลข ตัดช่องว่างในชื่อตำบล ลบแถวซ้ำ "
             "และตัดสินใจว่าจะจัดการค่า -999 อย่างไร พร้อมเขียนเหตุผลกำกับ"),
            ("Intermediate",
             "เชื่อมตารางพันธุ์เข้ากับข้อมูลแปลง แล้วสร้าง PivotTable เปรียบเทียบผลผลิตเฉลี่ยต่อไร่ "
             "ระหว่างพันธุ์ พร้อมแสดงจำนวนแปลงของแต่ละพันธุ์ควบคู่กัน",
             "ต้องใช้ IFERROR รองรับรหัสพันธุ์ที่ไม่มีในตารางอ้างอิง และต้องคำนวณผลผลิตต่อไร่เอง "
             "โดยระวังค่าพื้นที่ที่มีหน่วยติดอยู่ในเซลล์"),
            ("Advanced",
             "สร้าง Dashboard หนึ่งหน้าตอบคำถามว่า พันธุ์ใดให้กำไรต่อไร่สูงที่สุดเมื่อคิดต้นทุนแล้ว "
             "และใช้ AI ช่วยเขียนสูตรหรือสคริปต์ตรวจสอบคุณภาพข้อมูลอัตโนมัติ",
             "ต้องตรวจสอบผลลัพธ์จาก AI ด้วยการทดสอบกับแปลงที่คำนวณมือไว้ก่อน "
             "และต้องระบุข้อจำกัดของข้อสรุปเมื่อมีข้อมูลบางส่วนถูกตัดทิ้ง"),
        ],
        signature="ค่า -999 ที่ปนอยู่ในคอลัมน์ผลผลิต ถ้าเผลอนำไปเฉลี่ยจะทำให้ค่าเฉลี่ยติดลบหรือต่ำผิดปกติ"))

    # ---------------- AC ----------------
    rows, ref = DS.ac()
    # ทำให้บางใบสำคัญไม่ดุล — กับดักประจำสาขา
    vnos = sorted(set(r["เลขที่ใบสำคัญ"] for r in rows))
    broken = random.sample(vnos, 6)
    for v in broken:
        legs = [r for r in rows if r["เลขที่ใบสำคัญ"] == v]
        tgt = legs[0]
        key = "เดบิต" if tgt["เดบิต"] else "เครดิต"
        tgt[key] = DS.money(tgt[key] + random.choice([100, -100, 900, 1000, -450, 27]), 2)
        log("AC", "ใบสำคัญที่เดบิตไม่เท่าเครดิต", "เลขที่ {}".format(v))
    trap_text_numbers(rows, "เดบิต", 26, "AC")
    # วงเล็บแบบงบการเงินใส่ในคอลัมน์ยอดคงเหลือ ซึ่งค่าติดลบมีความหมายจริง
    # ไม่ใส่ในคอลัมน์เดบิตหรือเครดิต เพราะจะทำให้บทเรียนเรื่องใบสำคัญไม่ดุลปนกับบทเรียนเรื่องวงเล็บ
    idx = [i for i, r in enumerate(rows) if r["ยอดคงเหลือศูนย์ต้นทุน (บาท)"] < 0]
    for i in idx:
        v = rows[i]["ยอดคงเหลือศูนย์ต้นทุน (บาท)"]
        rows[i]["ยอดคงเหลือศูนย์ต้นทุน (บาท)"] = "({:,.2f})".format(abs(v))
        log("AC", "ยอดติดลบแสดงด้วยวงเล็บแบบงบการเงิน", "แถว {} คอลัมน์ ยอดคงเหลือศูนย์ต้นทุน".format(i + 2))
    trap_thai_date(rows, "วันที่", 30, "AC")
    trap_padded(rows, "รหัสบัญชี", 16, "AC")
    trap_orphan_key(rows, "รหัสบัญชี", 7, "AC", "509000")
    trap_blank(rows, "ศูนย์ต้นทุน", 14, "AC")
    trap_duplicate_rows(rows, 8, "AC")
    specs.append(dict(
        code="AC", prog="การบัญชี",
        fname="AC_สมุดรายวันทั่วไป.xlsx",
        title="สมุดรายวันทั่วไป ไตรมาสที่ 1–2 ปีบัญชี 2569 (ข้อมูลจำลอง)",
        context="รายการบันทึกบัญชีจากใบสำคัญทั่วไป ส่งออกมาจากระบบบัญชีเก่าที่ไม่ได้ควบคุมรูปแบบข้อมูล "
                "จึงมีทั้งจำนวนเงินที่เป็นข้อความ วงเล็บแสดงยอดติดลบ และรหัสบัญชีที่มีช่องว่างปน",
        rows=rows, refs=[ref], traps=8,
        tasks=[
            ("Beginner",
             "ทำให้ทุกจำนวนเงินเป็นตัวเลขที่คำนวณได้ แล้วหาผลรวมเดบิตและเครดิตทั้งสมุด",
             "ต้องแปลงข้อความเป็นตัวเลข และแปลงวงเล็บให้เป็นค่าติดลบ ซึ่งเป็นธรรมเนียมของงบการเงิน "
             "ไม่ใช่ข้อผิดพลาด จึงต้องเข้าใจความหมายก่อนแปลง"),
            ("Intermediate",
             "เชื่อมผังบัญชีเข้ามา แล้วหาว่าใบสำคัญเลขที่ใดบ้างที่เดบิตไม่เท่ากับเครดิต",
             "ใช้ SUMIFS จับคู่ตามเลขที่ใบสำคัญ แล้วเปรียบเทียบผลต่าง วิธีที่ดีคือสร้างคอลัมน์ผลต่าง "
             "แล้วกรองเฉพาะที่ไม่เป็นศูนย์ ระวังเรื่องการปัดเศษทศนิยม"),
            ("Advanced",
             "สร้างระบบตรวจสอบอัตโนมัติที่รายงานใบสำคัญผิดดุลทันทีเมื่อมีข้อมูลใหม่ "
             "และสรุปยอดตามประเภทบัญชีเป็นงบทดลอง",
             "ให้ AI ช่วยร่างสูตรหรือ Apps Script แล้วต้องทดสอบด้วยใบสำคัญที่ทราบคำตอบล่วงหน้า "
             "รวมถึงกรณีที่รหัสบัญชีไม่มีในผังบัญชี"),
        ],
        signature="มีใบสำคัญ 6 ใบที่เดบิตไม่เท่าเครดิต ซึ่งมองไม่เห็นถ้าไม่จับคู่ตามเลขที่ใบสำคัญ "
                  "และซ่อนตัวดีขึ้นอีกเพราะบางยอดถูกเก็บเป็นข้อความ"))

    # ---------------- BA ----------------
    rows, ref = DS.ba()
    trap_text_numbers(rows, "ราคาต่อหน่วย", 24, "BA")
    idx = random.sample(range(len(rows)), 18)
    for i in idx:
        rows[i]["ส่วนลด (บาท)"] = "{:.0f}%".format(random.choice([10, 15, 20]))
        log("BA", "ส่วนลดปนกันระหว่างบาทกับเปอร์เซ็นต์", "แถว {}".format(i + 2))
    idx = random.sample(range(len(rows)), 11)
    for i in idx:
        rows[i]["จำนวน"] = -rows[i]["จำนวน"]
        log("BA", "จำนวนติดลบจากรายการคืนสินค้า (ไม่ใช่ข้อผิดพลาด)", "แถว {}".format(i + 2))
    trap_variant_spelling(rows, "สาขา", {"สาขาศาลายา": "ศาลายา", "สาขากาญจนบุรี": "สาขา กาญจนบุรี",
                                          "สาขาท่าม่วง": "ท่าม่วง "}, 20, "BA")
    trap_orphan_key(rows, "รหัสสินค้า", 9, "BA", "BEV-099")
    trap_thai_date(rows, "วันที่", 28, "BA")
    trap_blank(rows, "รหัสพนักงาน", 13, "BA")
    trap_duplicate_rows(rows, 10, "BA")
    specs.append(dict(
        code="BA", prog="บริหารธุรกิจ",
        fname="BA_ยอดขายร้านกาแฟ.xlsx",
        title="รายการขายรายใบเสร็จ ร้านกาแฟ 4 สาขา มกราคม–พฤษภาคม 2569 (ข้อมูลจำลอง)",
        context="ข้อมูลส่งออกจากระบบ POS ที่แต่ละสาขาตั้งค่าไม่เหมือนกัน มีทั้งการคืนสินค้าที่บันทึกเป็นจำนวนติดลบ "
                "และส่วนลดที่บางสาขากรอกเป็นบาท บางสาขากรอกเป็นเปอร์เซ็นต์",
        rows=rows, refs=[ref], traps=8,
        tasks=[
            ("Beginner",
             "ทำความสะอาดชื่อสาขาให้เป็นมาตรฐานเดียวกัน แล้วหายอดขายรวมของแต่ละสาขา",
             "ชื่อสาขาที่สะกดต่างกันจะทำให้ COUNTIF และ PivotTable แยกเป็นคนละกลุ่ม "
             "ต้องตัดสินใจด้วยว่าจะรวมรายการคืนสินค้าเข้าไปในยอดขายหรือไม่ และให้เหตุผล"),
            ("Intermediate",
             "เชื่อมตารางสินค้าเพื่อคำนวณกำไรขั้นต้นรายรายการ แล้วสร้าง PivotTable "
             "หาว่าหมวดสินค้าใดทำกำไรรวมสูงที่สุดในแต่ละสาขา",
             "ต้องแก้ส่วนลดที่ปนกันสองหน่วยให้เป็นฐานเดียวกันก่อน และใช้ IFERROR "
             "รองรับรหัสสินค้าที่ไม่มีในตารางอ้างอิง"),
            ("Advanced",
             "สร้าง Dashboard ยอดขายที่ผู้จัดการเปิดดูเองได้ พร้อมข้อเสนอว่าควรทำโปรโมชันกับสินค้าใด "
             "ในช่วงเวลาใดของวัน",
             "ต้องแยกช่วงเวลาจากคอลัมน์เวลา และตรวจสอบว่าข้อสรุปไม่ได้มาจากกลุ่มที่มีจำนวนรายการน้อยเกินไป"),
        ],
        signature="ส่วนลดที่ปนกันระหว่างบาทกับเปอร์เซ็นต์ ถ้าเผลอบวกรวมกันจะได้ยอดที่ผิดโดยไม่มีสัญญาณเตือน"))

    # ---------------- CB ----------------
    rows, ref = DS.cb()
    idx = random.sample(range(len(rows)), 15)
    for i in idx:
        rows[i]["จำนวนตัว"] = DS.pick(["หลายตัว", "2-3", "ไม่แน่ชัด", "1 ตัว"])
        log("CB", "จำนวนถูกกรอกเป็นคำบรรยายแทนตัวเลข", "แถว {}".format(i + 2))
    trap_thai_date(rows, "วันที่", 24, "CB")
    trap_padded(rows, "รหัสชนิด", 17, "CB")
    trap_orphan_key(rows, "รหัสชนิด", 10, "CB", "SP-99")
    trap_blank(rows, "เพศ", 20, "CB")
    trap_outlier(rows, "อุณหภูมิ (°C)", 6, "CB", 99.9)
    trap_duplicate_rows(rows, 12, "CB")
    idx = random.sample(range(len(rows)), 9)
    for i in idx:
        rows[i]["เวลา"] = str(rows[i]["เวลา"]).replace(":", ".")
        log("CB", "รูปแบบเวลาไม่สม่ำเสมอ (ใช้จุดแทนทวิภาค)", "แถว {}".format(i + 2))
    specs.append(dict(
        code="CB", prog="ชีววิทยาเชิงอนุรักษ์",
        fname="CB_สำรวจกล้องดักถ่าย.xlsx",
        title="บันทึกภาพจากกล้องดักถ่ายสัตว์ป่า 7 จุดสำรวจ พฤศจิกายน 2568 – มิถุนายน 2569 (ข้อมูลจำลอง)",
        context="ข้อมูลจำแนกภาพจากกล้องดักถ่าย บันทึกโดยทีมสำรวจ อาสาสมัคร และนักศึกษาฝึกงาน "
                "ซึ่งได้รับการอบรมต่างระดับกัน จึงมีทั้งการกรอกจำนวนเป็นคำบรรยาย และการบันทึกซ้ำจากภาพเดียวกัน",
        rows=rows, refs=[ref], traps=8,
        tasks=[
            ("Beginner",
             "ทำความสะอาดข้อมูลแล้วนับจำนวนภาพที่พบของแต่ละชนิด พร้อมระบุจุดสำรวจที่พบชนิดมากที่สุด",
             "ต้องตัดสินใจว่าจะทำอย่างไรกับจำนวนที่กรอกเป็นคำบรรยาย ห้ามลบทิ้งเงียบ ๆ "
             "ให้แยกออกมาเป็นกลุ่มต่างหากพร้อมนับจำนวน และตัดช่องว่างในรหัสชนิดก่อนนับ"),
            ("Intermediate",
             "เชื่อมตารางชนิดเพื่อดึงชื่อไทยและสถานะการอนุรักษ์ แล้วสร้าง PivotTable "
             "แสดงความถี่การพบของชนิดที่มีสถานะเสี่ยง แยกตามจุดสำรวจและเดือน",
             "ต้องแยกเดือนออกจากวันที่ที่มีรูปแบบปนกัน และใช้ IFERROR กับรหัสชนิดที่ยังจำแนกไม่ได้"),
            ("Advanced",
             "คำนวณดัชนีความถี่การพบต่อคืนกล้อง (naive occupancy) และสร้างรายงานที่เสนอว่าควรเพิ่มกล้องที่จุดใด "
             "โดยใช้ AI ช่วยออกแบบสูตร",
             "ต้องระวังการนับซ้ำจากแถวที่ซ้ำกัน ซึ่งจะทำให้ดัชนีสูงเกินจริง และต้องระบุข้อจำกัดว่าดัชนีนี้ "
             "ไม่ได้ปรับค่าความน่าจะเป็นในการตรวจพบ"),
        ],
        signature="แถวซ้ำที่เกิดจากการบันทึกภาพเดียวกันสองครั้ง ซึ่งจะทำให้ความถี่การพบสูงเกินจริง "
                  "และเป็นข้อผิดพลาดที่ส่งผลตรงต่อข้อสรุปเชิงอนุรักษ์"))

    # ---------------- ED ----------------
    rows, ref, ref2 = DS.ed()
    trap_sentinel(rows, "PM2.5 (µg/m³)", 34, "ED")
    trap_sentinel(rows, "ปริมาณน้ำฝน (มม.)", 20, "ED")
    trap_outlier(rows, "PM10 (µg/m³)", 12, "ED", 9999)
    trap_blank(rows, "ระดับน้ำ (ม.รทก.)", 26, "ED")
    trap_text_numbers(rows, "อุณหภูมิ (°C)", 35, "ED")
    trap_thai_date(rows, "วันที่", 45, "ED")
    trap_padded(rows, "รหัสสถานี", 24, "ED")
    # ลบข้อมูลออกทั้งวัน (ทุกสถานี) เพื่อสร้างช่องว่างของอนุกรมเวลาที่ตรวจพบได้จริง
    all_days = sorted(set(str(r["วันที่"]) for r in rows))
    drop_days = set(random.sample(all_days[5:-5], 6))
    for dd in sorted(drop_days):
        log("ED", "วันที่หายไปทั้งวันจากอนุกรมเวลา", "วันที่ {}".format(dd))
    rows[:] = [r for r in rows if str(r["วันที่"]) not in drop_days]
    # และลบเฉพาะบางสถานีในบางวัน เลียนแบบสถานีเดี่ยวที่ส่งข้อมูลไม่สำเร็จ
    partial = random.sample(range(len(rows)), 30)
    for i in sorted(partial, reverse=True):
        log("ED", "ข้อมูลของบางสถานีหายไปเฉพาะบางวัน", "แถว {}".format(i + 2))
        rows.pop(i)
    specs.append(dict(
        code="ED", prog="วิศวกรรมสิ่งแวดล้อมและการจัดการภัยพิบัติ",
        fname="ED_สถานีตรวจวัดสิ่งแวดล้อม.xlsx",
        title="ข้อมูลตรวจวัดคุณภาพอากาศ ปริมาณน้ำฝน และระดับน้ำ 6 สถานี กันยายน 2568 – มกราคม 2569 (ข้อมูลจำลอง)",
        context="ข้อมูลดิบจากสถานีตรวจวัดอัตโนมัติ ซึ่งใช้ค่า -999 แทนช่วงที่เซนเซอร์ไม่ทำงาน "
                "และมีบางวันที่ข้อมูลหายไปทั้งแถวเพราะระบบส่งข้อมูลล่ม",
        rows=rows, refs=[ref, ref2], traps=8,
        tasks=[
            ("Beginner",
             "จัดการค่า -999 และค่าผิดปกติให้เรียบร้อย แล้วหาค่าเฉลี่ย PM2.5 ของแต่ละสถานี",
             "ห้ามใช้ค่าเฉลี่ยจากข้อมูลดิบเด็ดขาด เพราะ -999 จะดึงค่าเฉลี่ยลงจนติดลบ "
             "ต้องเปลี่ยนเป็นค่าว่างหรือกรองออก แล้วรายงานจำนวนวันที่ใช้คำนวณจริงควบคู่ไปด้วย"),
            ("Intermediate",
             "เชื่อมตารางสถานีและเกณฑ์มาตรฐาน แล้วนับจำนวนวันที่แต่ละสถานีมีค่าเกินเกณฑ์ "
             "พร้อมเปรียบเทียบระหว่างพื้นที่เมือง เกษตร และป่า",
             "ต้องใช้ COUNTIFS ที่ไม่นับค่าที่ถูกตัดออก และต้องระวังว่าจำนวนวันที่มีข้อมูลของแต่ละสถานีไม่เท่ากัน "
             "จึงต้องรายงานเป็นสัดส่วน ไม่ใช่จำนวนวันดิบ"),
            ("Advanced",
             "สร้างระบบเฝ้าระวังที่คำนวณค่าเฉลี่ยเคลื่อนที่ 7 วัน และแจ้งเตือนเมื่อมีแนวโน้มเกินเกณฑ์ "
             "พร้อม Dashboard เปรียบเทียบสถานี",
             "ค่าเฉลี่ยเคลื่อนที่จะผิดทันทีถ้าอนุกรมเวลามีวันขาดหาย ต้องสร้างปฏิทินวันที่ให้ครบก่อน "
             "แล้วจึงเชื่อมข้อมูลเข้ามา ซึ่งเป็นบทเรียนสำคัญของงานอนุกรมเวลา"),
        ],
        signature="วันที่หายไปจากอนุกรมเวลา 22 วัน ซึ่งไม่มีอะไรบอกให้รู้เลยถ้าไม่ตรวจ "
                  "และทำให้ค่าเฉลี่ยเคลื่อนที่คลาดเคลื่อนโดยไม่มีข้อความ error ใด ๆ"))

    # ---------------- FT ----------------
    rows, ref = DS.ft()
    idx = random.sample(range(len(rows)), 26)
    for i in idx:
        rows[i]["จุลินทรีย์ทั้งหมด (CFU/g)"] = DS.pick(["<10", "ND", "<100", "TNTC", "ไม่พบ"])
        log("FT", "ผลตรวจจุลินทรีย์เป็นข้อความตามธรรมเนียมห้องปฏิบัติการ", "แถว {}".format(i + 2))
    idx = random.sample(range(len(rows)), 12)
    for i in idx:
        rows[i]["น้ำหนักสุทธิ (ก.)"] = DS.money(rows[i]["น้ำหนักสุทธิ (ก.)"] / 1000, 3)
        log("FT", "น้ำหนักบางแถวบันทึกเป็นกิโลกรัมแทนกรัม", "แถว {}".format(i + 2))
    trap_text_numbers(rows, "pH", 18, "FT")
    trap_thai_date(rows, "วันที่ผลิต", 22, "FT")
    trap_blank(rows, "aw", 15, "FT")
    trap_padded(rows, "รหัสผลิตภัณฑ์", 13, "FT")
    trap_orphan_key(rows, "รหัสผลิตภัณฑ์", 6, "FT", "FP-99")
    trap_duplicate_rows(rows, 9, "FT")
    specs.append(dict(
        code="FT", prog="เทคโนโลยีการอาหาร",
        fname="FT_บันทึกควบคุมคุณภาพสายการผลิต.xlsx",
        title="บันทึกควบคุมคุณภาพระหว่างการผลิต 380 ล็อต กุมภาพันธ์–มิถุนายน 2569 (ข้อมูลจำลอง)",
        context="บันทึก QC จากสายการผลิต 3 สาย ที่ผู้ตรวจแต่ละกะกรอกด้วยมือ "
                "ผลตรวจจุลินทรีย์ใช้ธรรมเนียมห้องปฏิบัติการ เช่น <10 หรือ ND ซึ่งไม่ใช่ตัวเลขที่คำนวณได้ทันที",
        rows=rows, refs=[ref], traps=8,
        tasks=[
            ("Beginner",
             "ทำให้ค่า pH และน้ำหนักสุทธิเป็นตัวเลขที่เปรียบเทียบกันได้ แล้วหาค่าเฉลี่ยรายผลิตภัณฑ์",
             "ต้องสังเกตว่าน้ำหนักบางแถวเป็นกิโลกรัม การเฉลี่ยรวมโดยไม่แปลงหน่วยจะดึงค่าเฉลี่ยลงมาก "
             "และต้องอธิบายวิธีจัดการผลตรวจที่เป็นข้อความอย่างมีหลักการ"),
            ("Intermediate",
             "เชื่อมสเปกผลิตภัณฑ์เข้ามา แล้วหาว่าล็อตใดบ้างที่หลุดสเปก pH หรือน้ำหนัก "
             "พร้อมสรุปว่ากะใดและสายการผลิตใดมีปัญหามากที่สุด",
             "ต้องใช้เงื่อนไขหลายชั้นเทียบกับช่วงต่ำสุด–สูงสุดที่ต่างกันในแต่ละผลิตภัณฑ์ "
             "และใช้ IFERROR กับรหัสผลิตภัณฑ์ที่ไม่มีในสเปก"),
            ("Advanced",
             "สร้างแผนภูมิควบคุม (control chart) ของน้ำหนักสุทธิ พร้อมระบบธงเตือนล็อตที่หลุดสเปก "
             "และร่างขั้นตอนการตรวจสอบย้อนกลับเมื่อพบปัญหา",
             "ต้องตัดสินใจว่าจะแปลง <10 เป็นค่าใดในการคำนวณ ซึ่งเป็นการตัดสินใจเชิงวิชาการที่ต้องบันทึกเหตุผลไว้ "
             "และให้ AI ช่วยได้แต่ต้องตรวจสอบเอง"),
        ],
        signature="ผลตรวจจุลินทรีย์ที่เป็น <10, ND, TNTC ซึ่งเป็นค่าที่ถูกต้องตามธรรมเนียมห้องปฏิบัติการ "
                  "แต่คำนวณตรง ๆ ไม่ได้ ผู้เรียนต้องตัดสินใจเชิงวิชาการ ไม่ใช่แค่ลบทิ้ง"))

    # ---------------- GS ----------------
    rows, ref = DS.gs()
    # สลับความลึกบนกับล่างในบางแถว — ผิดตรรกะแต่ไม่มี error
    idx = random.sample(range(len(rows)), 13)
    for i in idx:
        rows[i]["ความลึกช่วงบน (ม.)"], rows[i]["ความลึกช่วงล่าง (ม.)"] = \
            rows[i]["ความลึกช่วงล่าง (ม.)"], rows[i]["ความลึกช่วงบน (ม.)"]
        log("GS", "ความลึกช่วงบนมากกว่าช่วงล่าง (ผิดตรรกะ)", "แถว {}".format(i + 2))
    idx = random.sample(range(len(rows)), 20)
    for i in idx:
        rows[i]["พิกัด E"] = "{:,}".format(rows[i]["พิกัด E"])
        log("GS", "พิกัดถูกเก็บเป็นข้อความพร้อมเครื่องหมายคั่นหลักพัน", "แถว {}".format(i + 2))
    trap_outlier(rows, "TDS (mg/L)", 8, "GS", 99999)
    trap_blank(rows, "pH น้ำบาดาล", 14, "GS")
    trap_variant_spelling(rows, "รหัสหน่วยหิน", {"Qa": "QA", "Tr-ss": "TR-SS", "Pm-ls": "Pm-Ls"}, 18, "GS")
    trap_thai_date(rows, "วันที่เจาะ", 21, "GS")
    trap_sentinel(rows, "ปริมาณน้ำ (ลบ.ม./ชม.)", 9, "GS", 0)
    trap_duplicate_rows(rows, 7, "GS")
    specs.append(dict(
        code="GS", prog="ธรณีศาสตร์",
        fname="GS_ข้อมูลเจาะสำรวจน้ำบาดาล.xlsx",
        title="ข้อมูลหลุมเจาะสำรวจน้ำบาดาล 290 หลุม มิถุนายน 2568 – เมษายน 2569 (ข้อมูลจำลอง)",
        context="ข้อมูลรวบรวมจากรายงานหลุมเจาะของทีมเจาะสามทีมและผู้รับเหมา ซึ่งใช้รูปแบบการเขียนรหัสหน่วยหิน "
                "และการบันทึกพิกัดต่างกัน จึงต้องปรับให้เป็นมาตรฐานเดียวกันก่อนวิเคราะห์",
        rows=rows, refs=[ref], traps=7,
        tasks=[
            ("Beginner",
             "ทำให้พิกัดและค่าวิเคราะห์เป็นตัวเลขที่ใช้งานได้ แล้วหาค่าเฉลี่ย TDS ของแต่ละหน่วยหิน",
             "ต้องปรับตัวพิมพ์ใหญ่เล็กของรหัสหน่วยหินให้ตรงกันก่อน มิฉะนั้นจะได้กลุ่มซ้ำซ้อน "
             "และต้องจัดการค่า TDS ที่สูงผิดปกติอย่างมีเหตุผล"),
            ("Intermediate",
             "เชื่อมตารางหน่วยหิน แล้วสร้าง PivotTable เปรียบเทียบปริมาณน้ำเฉลี่ยและความลึกเฉลี่ย "
             "ระหว่างหน่วยหินที่มีศักยภาพต่างกัน",
             "ต้องคำนวณความหนาของชั้นจากความลึกบนและล่าง ซึ่งจะเผยให้เห็นแถวที่ค่าติดลบ "
             "อันเป็นสัญญาณว่าข้อมูลถูกบันทึกสลับกัน"),
            ("Advanced",
             "สร้างเกณฑ์คัดกรองพื้นที่ที่เหมาะสมสำหรับเจาะบ่อใหม่ โดยพิจารณาศักยภาพน้ำ คุณภาพน้ำ และความลึก "
             "พร้อมแผนภาพสรุปเชิงพื้นที่",
             "ต้องตั้งเกณฑ์ให้ชัดและอธิบายเหตุผลของแต่ละเกณฑ์ ให้ AI ช่วยร่างสูตรคัดกรองได้ "
             "แต่ต้องตรวจสอบกับหลุมที่ทราบผลจริงก่อนนำไปใช้"),
        ],
        signature="แถวที่ความลึกช่วงบนมากกว่าช่วงล่าง ซึ่งเป็นไปไม่ได้ทางกายภาพ แต่ Excel ไม่เตือน "
                  "จะพบก็ต่อเมื่อคำนวณความหนาแล้วเห็นค่าติดลบ"))

    return specs


# ============================================================ เขียนไฟล์
def style_header(ws, ncols, fill=HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"


def write_book(spec, outdir):
    wb = Workbook()

    # ---- ชีตอ่านก่อนเริ่ม ----
    ws = wb.active
    ws.title = "อ่านก่อนเริ่ม"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 104
    r = 2
    def put(label, text, bold=False, color=NAVY, size=10):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, bold=True, size=size, color=color)
        c = ws.cell(row=r, column=3, value=text)
        c.font = Font(name=FONT, size=size, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if text and len(text) > 100:
            ws.row_dimensions[r].height = 15 * (len(text) // 100 + 1)
        r += 1

    ws.cell(row=r, column=2, value=spec["code"]).font = Font(name=FONT, bold=True, size=22, color=NAVY)
    ws.cell(row=r, column=3, value=spec["title"]).font = Font(name=FONT, bold=True, size=14, color=NAVY)
    r += 1
    put("", "ชุดการเรียนรู้ Excel & Google Sheets · สาขา" + spec["prog"] + " · อาจารย์ ดร.เนติยา การะเกตุ")
    r += 1
    put("บริบทของข้อมูล", spec["context"])
    put("ที่มาของข้อมูล", "ข้อมูลทั้งหมดในไฟล์นี้เป็นข้อมูลจำลองที่สร้างขึ้นด้วยโปรแกรมเพื่อการเรียนการสอนเท่านั้น "
                          "ไม่ใช่ข้อมูลจริงของบุคคล หน่วยงาน หรือสถานประกอบการใด "
                          "ห้ามนำตัวเลขในไฟล์นี้ไปอ้างอิงในงานวิชาการหรือรายงานใด ๆ")
    put("จำนวนแถวข้อมูล", "{:,} แถว ในชีต “ข้อมูลดิบ”".format(len(spec["rows"])))
    r += 1

    ws.cell(row=r, column=2, value="ความท้าทาย").font = Font(name=FONT, bold=True, size=13, color="C00000")
    ws.cell(row=r, column=3, value="ไฟล์นี้มีปัญหาข้อมูลซ่อนอยู่ {} ประเภท".format(spec["traps"])).font = \
        Font(name=FONT, bold=True, size=13, color="C00000")
    r += 1
    put("", "ทั้งหมดเป็นปัญหาที่พบจริงในงานของสาขานี้ ไม่ใช่ความผิดพลาดที่ตั้งใจแกล้ง "
            "งานแรกของท่านคือหาให้ครบว่ามีปัญหาอะไรบ้าง แล้วบันทึกไว้ว่าจัดการแต่ละอย่างอย่างไรและเพราะเหตุใด "
            "การบันทึกเหตุผลสำคัญไม่แพ้การแก้ เพราะเป็นสิ่งที่ทำให้งานวิเคราะห์ตรวจสอบย้อนกลับได้")
    put("", "คำเตือน: อย่าลบข้อมูลที่ดูแปลกทิ้งทันที บางค่าที่ดูเหมือนผิดคือข้อมูลที่ถูกต้องตามธรรมเนียมของสาขา")
    r += 1

    ws.cell(row=r, column=2, value="โจทย์").font = Font(name=FONT, bold=True, size=13, color=NAVY)
    r += 1
    for lv, task, hint in spec["tasks"]:
        ws.cell(row=r, column=2, value=lv).font = Font(name=FONT, bold=True, size=11, color="117864")
        c = ws.cell(row=r, column=3, value=task)
        c.font = Font(name=FONT, size=10, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 15 * (len(task) // 100 + 1)
        r += 1
        c = ws.cell(row=r, column=3, value="ข้อควรระวัง: " + hint)
        c.font = Font(name=FONT, size=10, italic=True, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 15 * (len(hint) // 100 + 1)
        r += 2

    put("การส่งงาน", "ส่งไฟล์ที่ทำเสร็จผ่าน Google Classroom ของระดับที่ท่านลงทะเบียน "
                     "พร้อมชีตชื่อ “บันทึกการตัดสินใจ” ที่ระบุว่าท่านพบปัญหาอะไรและจัดการอย่างไร")

    for row in ws.iter_rows(min_row=1, max_row=r, min_col=2, max_col=3):
        for cell in row:
            if cell.value and not cell.font.bold:
                pass
    ws.sheet_view.showGridLines = False

    # ---- ชีตข้อมูลดิบ ----
    d = wb.create_sheet("ข้อมูลดิบ")
    cols = list(spec["rows"][0].keys())
    for i, c in enumerate(cols, 1):
        d.cell(row=1, column=i, value=c)
    style_header(d, len(cols))
    for ri, rec in enumerate(spec["rows"], start=2):
        for ci, c in enumerate(cols, 1):
            cell = d.cell(row=ri, column=ci)
            cell.value = rec[c]
            cell.font = Font(name=FONT, size=10)
            if isinstance(rec[c], str) and rec[c].startswith("="):
                cell.data_type = "s"
    for i, c in enumerate(cols, 1):
        d.column_dimensions[get_column_letter(i)].width = max(11, min(26, len(c) + 4))
    d.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(cols)), len(spec["rows"]) + 1)

    # ---- ชีตอ้างอิง ----
    for ref in spec["refs"]:
        s = wb.create_sheet(ref["sheet"])
        for i, c in enumerate(ref["cols"], 1):
            s.cell(row=1, column=i, value=c)
        style_header(s, len(ref["cols"]), GREEN)
        for ri, rec in enumerate(ref["rows"], start=2):
            for ci, v in enumerate(rec, 1):
                cell = s.cell(row=ri, column=ci, value=v)
                cell.font = Font(name=FONT, size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, c in enumerate(ref["cols"], 1):
            s.column_dimensions[get_column_letter(i)].width = max(14, min(34, len(str(c)) + 8))

    path = os.path.join(outdir, spec["fname"])
    wb.save(path)
    return path


if __name__ == "__main__":
    specs = build_all()
    # จำนวนกับดักที่ประกาศในไฟล์ต้องมาจากบันทึกจริง ไม่ใช่ตัวเลขที่พิมพ์มือ จะได้ไม่มีทางคลาดเคลื่อน
    for sp in specs:
        sp["traps"] = len(LOG.get(sp["code"], {}))
    made = []
    for sp in specs:
        made.append(write_book(sp, OUT))
        print("saved", sp["fname"], "|", len(sp["rows"]), "rows |", sp["traps"], "traps")
    meta = [{k: v for k, v in sp.items() if k != "rows"} for sp in specs]
    for m in meta:
        m["nrows"] = next(len(s["rows"]) for s in specs if s["code"] == m["code"])
        m["refs"] = [{"sheet": r["sheet"], "cols": r["cols"], "nrows": len(r["rows"])} for r in m["refs"]]
        m["trap_detail"] = {k: len(v) for k, v in LOG.get(m["code"], {}).items()}
    json.dump(meta, open("samples_meta.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump({k: {kk: vv for kk, vv in v.items()} for k, v in LOG.items()},
              open("samples_trap_log.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print("\nสรุปกับดักที่ฝังจริง")
    for code, traps in LOG.items():
        print(" ", code, "->", len(traps), "ประเภท,", sum(len(v) for v in traps.values()), "จุด")
