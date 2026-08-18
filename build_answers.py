#!/usr/bin/env python3
"""
สร้าง "ฉบับเฉลย" ของไฟล์ฝึกปฏิบัติทั้งเจ็ดสาขา
หลักการเดียวกับ answer_key.py คือทำความสะอาดจากไฟล์ที่ส่งมอบจริง ไม่ใช่จากบันทึกตอนสร้าง
เฉลยจึงตรงกับสิ่งที่ผู้เรียนเห็นเสมอ แม้ไฟล์ต้นทางจะถูกสร้างใหม่
"""
import json, os, re, sys, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1]
OUT = sys.argv[2]
CSVDIR = sys.argv[3]
os.makedirs(OUT, exist_ok=True)
os.makedirs(CSVDIR, exist_ok=True)

FONT = "Tahoma"
NAVY = "1F3864"; HEAD = "D9E2F3"; GREEN = "E2EFDA"; ORANGE = "FBE5D6"; GREY = "F2F2F2"
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

meta = {m["code"]: m for m in json.load(open("samples_meta.json", encoding="utf-8"))}
KEY = json.load(open("answer_key.json", encoding="utf-8"))

NUM_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")
BE_RE = re.compile(r"^(\d{2})/(\d{2})/(25\d{2})$")
UNIT_RE = re.compile(r"^(-?[\d,]+(?:\.\d+)?)\s*([^\d\s].*)$")
ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
# วงเล็บในงานบัญชีหมายถึงจำนวนลบ ไม่ใช่ข้อความ ต้องแปลงเป็นตัวเลขติดลบ
PAREN_RE = re.compile(r"^\(\s*([\d,]+(?:\.\d+)?)\s*\)$")
# คอลัมน์ที่เป็น "ตัวระบุ" ไม่ใช่ปริมาณ ห้ามแปลงเป็นตัวเลขเด็ดขาด
ID_RE = re.compile(r"รหัส|เลขที่|เบอร์|ไปรษณีย์|หมายเลข|\bID\b|code", re.IGNORECASE)
SENTINEL = {-999, -9999, 999999}


# ============================================================ กฎการทำความสะอาด
def clean_cell(v, stats, colname):
    """คืนค่าที่สะอาดแล้ว พร้อมนับว่าแก้ด้วยกฎใด"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v in SENTINEL:
            stats["ค่าแทนการไม่มีข้อมูลถูกเปลี่ยนเป็นช่องว่าง"] += 1
            return None
        return v
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    s = str(v)
    t = s.strip().rstrip(".").strip()
    if t != s.strip():
        stats["ตัดจุดหรืออักขระส่วนเกินท้ายข้อความ"] += 1
    if s != s.strip():
        stats["ตัดช่องว่างหน้าหลัง"] += 1
    t = re.sub(r"\s+", " ", t)
    if t == "":
        return None
    m = ISO_RE.match(t)
    if m:
        stats["แปลงวันที่แบบข้อความเป็นวันที่จริง"] += 1
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return t
    if colname and ID_RE.search(str(colname)):
        # รหัสหรือเลขที่เอกสารต้องคงเป็นข้อความ ถ้าแปลงเป็นตัวเลขจะเสียเลขศูนย์นำหน้า
        # และจะเชื่อมกับตารางอ้างอิงที่เก็บเป็นข้อความไม่ได้
        if NUM_RE.match(t):
            stats["คงรหัสไว้เป็นข้อความโดยตั้งใจ ไม่แปลงเป็นตัวเลข"] += 1
        return t
    m = BE_RE.match(t)
    if m:
        d, mo, by = int(m.group(1)), int(m.group(2)), int(m.group(3))
        stats["แปลงวันที่ พ.ศ. แบบข้อความเป็นวันที่จริง"] += 1
        try:
            return datetime.date(by - 543, mo, d)
        except ValueError:
            return t
    m = PAREN_RE.match(t)
    if m:
        stats["แปลงจำนวนในวงเล็บแบบงานบัญชีเป็นตัวเลขติดลบ"] += 1
        n = -float(m.group(1).replace(",", ""))
        return int(n) if n == int(n) else n
    if NUM_RE.match(t):
        if re.match(r"^0\d", t):
            stats["คงตัวเลขที่มีศูนย์นำหน้าไว้เป็นข้อความ"] += 1
            return t
        n = float(t.replace(",", ""))
        if "," in t:
            stats["แปลงตัวเลขที่มีเครื่องหมายคั่นหลักพัน"] += 1
        else:
            stats["แปลงตัวเลขที่เก็บเป็นข้อความ"] += 1
        if n in SENTINEL:
            stats["ค่าแทนการไม่มีข้อมูลถูกเปลี่ยนเป็นช่องว่าง"] += 1
            return None
        return int(n) if n == int(n) else n
    m = UNIT_RE.match(t)
    if m and colname and re.search(r"\(|%|กก|ซม|มม|บาท|ระดับ", str(colname)):
        stats["แยกหน่วยออกจากเซลล์ เหลือเฉพาะตัวเลข"] += 1
        n = float(m.group(1).replace(",", ""))
        return int(n) if n == int(n) else n
    return t


def normalize_categories(rows, hdr, stats):
    """รวมชื่อกลุ่มที่สะกดต่างกันเล็กน้อยให้เป็นรูปแบบที่พบบ่อยที่สุด"""
    for j, name in enumerate(hdr):
        vals = [r[j] for r in rows if isinstance(r[j], str)]
        if len(vals) < len(rows) * 0.5 or not vals:
            continue
        groups = {}
        for v in vals:
            # ใช้ตัวพิมพ์เล็กเป็นกุญแจด้วย เพราะ SUMIFS และ COUNTIFS ของ Excel ไม่แยกตัวพิมพ์
            # ถ้าปล่อยให้มีทั้ง Pm-ls และ Pm-Ls ตารางสรุปจะนับข้อมูลชุดเดียวกันสองครั้ง
            # แล้วผลรวมของทุกกลุ่มจะเกินผลรวมทั้งหมด ซึ่งตรวจพบได้จากช่องร้อยละที่เกิน 100
            k = re.sub(r"[\s.]", "", v).lower()
            groups.setdefault(k, []).append(v)
        if len(groups) == len(set(vals)):
            continue          # ไม่มีตัวไหนที่ต่างกันแค่ช่องว่างหรือจุด
        canon = {}
        for k, vs in groups.items():
            best = max(set(vs), key=vs.count)
            for v in set(vs):
                if v != best:
                    canon[v] = best
        if canon:
            n = 0
            for r in rows:
                if isinstance(r[j], str) and r[j] in canon:
                    r[j] = canon[r[j]]
                    n += 1
            if n:
                case_only = all(a.lower() == b.lower() for a, b in canon.items())
                label = ("รวมชื่อกลุ่มที่ต่างกันเพียงตัวพิมพ์ใหญ่เล็กในคอลัมน์ "
                         if case_only else "รวมชื่อกลุ่มที่สะกดต่างกันในคอลัมน์ ")
                stats[label + str(name)] += n
    return rows



def split_unparseable(hdr, rows, stats):
    """คอลัมน์ที่ตั้งใจเก็บตัวเลข แต่มีข้อความบรรยายปนอยู่ เช่น หลายตัว หรือ 2-3
    ข้อความเหล่านี้แปลงเป็นตัวเลขไม่ได้และไม่ควรเดาแทนผู้บันทึก
    วิธีที่ถูกคือย้ายไปเก็บในคอลัมน์หมายเหตุ แล้วปล่อยช่องตัวเลขให้ว่าง
    ข้อมูลเดิมจึงไม่สูญหาย และคอลัมน์ตัวเลขก็คำนวณได้ทั้งคอลัมน์"""
    add = []
    for j, name in enumerate(hdr):
        vals = [r[j] for r in rows if r[j] is not None and r[j] != ""]
        if not vals:
            continue
        nums = [v for v in vals if isinstance(v, (int, float))]
        txts = [v for v in vals if isinstance(v, str)]
        if not txts or len(nums) < len(vals) * 0.8:
            continue
        add.append((j, name, len(txts)))
    for j, name, n in add:
        note = str(name) + " หมายเหตุ"
        hdr.append(note)
        for r in rows:
            r.append(None)
        for r in rows:
            if isinstance(r[j], str) and r[j].strip() != "":
                r[-1] = r[j]
                r[j] = None
        stats["ย้ายข้อความที่แปลงเป็นตัวเลขไม่ได้ไปคอลัมน์ " + note] += n
    return hdr, rows


def clean_sheet(hdr, raw):
    from collections import Counter
    stats = Counter()
    rows = []
    for r in raw:
        rows.append([clean_cell(v, stats, hdr[j] if j < len(hdr) else None)
                     for j, v in enumerate(r)])
    rows = normalize_categories(rows, hdr, stats)
    hdr, rows = split_unparseable(hdr, rows, stats)
    seen, out = set(), []
    for r in rows:
        k = tuple("" if v is None else str(v) for v in r)
        if all(v == "" for v in k):
            stats["ลบแถวว่างทั้งแถว"] += 1
            continue
        if k in seen:
            stats["ลบแถวที่ซ้ำกันทั้งแถว"] += 1
            continue
        seen.add(k)
        out.append(r)
    return out, stats


# ============================================================ ตารางสรุปแบบ Pivot
# เลือกคอลัมน์หมวดและคอลัมน์ตัวเลขที่ตอบโจทย์ของสาขานั้นจริง ไม่ได้เลือกโดยอัตโนมัติ
# เพราะตารางสรุปที่ดีต้องตอบคำถามของสาขา ไม่ใช่แค่จับคอลัมน์แรกที่เจอ
PIVOT = {
    "AG": ("ตำบล", "ผลผลิต (กก.)", "ตำบลใดมีผลผลิตรวมสูงที่สุด และค่าเฉลี่ยต่อแปลงเป็นเท่าไร"),
    "AC": ("ศูนย์ต้นทุน", "เดบิต", "ศูนย์ต้นทุนใดมีรายการเดบิตรวมสูงที่สุด และมีกี่รายการ"),
    "BA": ("สาขา", "จำนวน", "สาขาใดขายได้จำนวนหน่วยรวมสูงที่สุด และเฉลี่ยต่อบิลเท่าไร"),
    "CB": ("รหัสจุดสำรวจ", "จำนวนตัว", "จุดสำรวจใดพบสัตว์รวมมากที่สุด และมีกี่ครั้งที่บันทึก"),
    "ED": ("รหัสสถานี", "PM2.5 (µg/m³)", "สถานีใดมีค่า PM2.5 เฉลี่ยสูงที่สุด และมีข้อมูลกี่วัน"),
    "FT": ("สายการผลิต", "pH", "สายการผลิตใดมีค่า pH เฉลี่ยเบี่ยงจากสเปกมากที่สุด"),
    "GS": ("รหัสหน่วยหิน", "ปริมาณน้ำ (ลบ.ม./ชม.)", "หน่วยหินใดให้ปริมาณน้ำเฉลี่ยสูงที่สุด และเจาะไปกี่บ่อ"),
}


def col_letter_of(hdr, name):
    return get_column_letter(hdr.index(name) + 1)


def add_pivot_sheets(wb, code, hdr, clean):
    cat, num, question = PIVOT[code]
    if cat not in hdr or num not in hdr:
        return None
    ci, ni = col_letter_of(hdr, cat), col_letter_of(hdr, num)
    last = len(clean) + 1
    D = "'ข้อมูลสะอาด'!"
    crange = "{d}${c}$2:${c}${n}".format(d=D, c=ci, n=last)
    nrange = "{d}${c}$2:${c}${n}".format(d=D, c=ni, n=last)

    groups = sorted(set(str(r[hdr.index(cat)]) for r in clean if r[hdr.index(cat)] not in (None, "")))

    ws = wb.create_sheet("ตารางสรุปแบบ Pivot")
    ws["A1"] = "ตารางสรุปแบบ Pivot ที่คำนวณสดด้วยสูตร"
    ws["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws["A2"] = ("โจทย์ที่ตารางนี้ตอบ: " + question)
    ws["A2"].font = Font(name=FONT, size=10, color="333333")
    ws["A3"] = ("ตารางนี้ไม่ใช่ PivotTable แต่ให้ผลลัพธ์เหมือนกันทุกช่อง ข้อดีคือเห็นสูตรที่อยู่เบื้องหลัง "
                "และแก้ค่าในชีตข้อมูลสะอาดแล้วตัวเลขที่นี่เปลี่ยนตามทันทีโดยไม่ต้องกดรีเฟรช "
                "ให้ผู้เรียนทำ PivotTable จริงตามขั้นตอนในชีตถัดไป แล้วเทียบกับตารางนี้ว่าตรงกันหรือไม่")
    ws["A3"].font = Font(name=FONT, size=10, color="555555")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 46
    ws.merge_cells("A3:F3")

    head = [cat, "จำนวนรายการ", "ผลรวม " + num, "ค่าเฉลี่ย " + num,
            "ผลรวมยืนยันด้วยวิธีที่สอง", "ร้อยละของผลรวมทั้งหมด"]
    for j, h in enumerate(head, start=1):
        ws.cell(row=5, column=j, value=h)
    style_head(ws, len(head), row=5, fill=GREEN)

    r = 6
    for g in groups:
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value="=COUNTIFS({c},$A{r})".format(c=crange, r=r))
        ws.cell(row=r, column=3, value="=SUMIFS({n},{c},$A{r})".format(n=nrange, c=crange, r=r))
        ws.cell(row=r, column=4, value=("=IFERROR(AVERAGEIFS({n},{c},$A{r}),"
                                        "\"ไม่มีข้อมูล\")").format(n=nrange, c=crange, r=r))
        ws.cell(row=r, column=5, value=("=IFERROR(SUMPRODUCT(({c}=$A{r})*{n}),"
                                        "\"พบข้อความปนในคอลัมน์ตัวเลข\")").format(
                                            n=nrange, c=crange, r=r))
        ws.cell(row=r, column=6, value="=IFERROR($C{r}/SUM({n}),0)".format(r=r, n=nrange))
        ws.cell(row=r, column=6).number_format = "0.0%"
        r += 1

    blanks = sum(1 for x in clean if x[hdr.index(cat)] in (None, ""))
    if blanks:
        ws.cell(row=r, column=1, value="ไม่ได้ระบุ")
        ws.cell(row=r, column=2, value='=COUNTIFS({c},"")'.format(c=crange))
        ws.cell(row=r, column=3, value='=SUMIFS({n},{c},"")'.format(n=nrange, c=crange))
        ws.cell(row=r, column=4, value='=IFERROR(AVERAGEIFS({n},{c},""),"ไม่มีข้อมูล")'.format(
            n=nrange, c=crange))
        ws.cell(row=r, column=5, value='=IFERROR(SUMPRODUCT(({c}="")*{n}),"พบข้อความปนในคอลัมน์ตัวเลข")'.format(
            n=nrange, c=crange))
        ws.cell(row=r, column=6, value="=IFERROR($C{r}/SUM({n}),0)".format(r=r, n=nrange))
        ws.cell(row=r, column=6).number_format = "0.0%"
        for j in range(1, 7):
            ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FBE5D6")
        r += 1

    ws.cell(row=r, column=1, value="รวมทั้งหมด").font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.cell(row=r, column=2, value="=SUM(B6:B{})".format(r - 1))
    ws.cell(row=r, column=3, value="=SUM(C6:C{})".format(r - 1))
    ws.cell(row=r, column=4, value="=IFERROR(C{r}/B{r},\"ไม่มีข้อมูล\")".format(r=r))
    ws.cell(row=r, column=5, value="=SUM(E6:E{})".format(r - 1))
    ws.cell(row=r, column=6, value="=SUM(F6:F{})".format(r - 1))
    ws.cell(row=r, column=6).number_format = "0.0%"
    for j in range(1, len(head) + 1):
        ws.cell(row=r, column=j).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=ORANGE)
    total_row = r

    for row in ws.iter_rows(min_row=6, max_row=total_row):
        for c in row:
            if c.font.bold is not True:
                c.font = Font(name=FONT, size=10)
            c.border = BORDER
    autowidth(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 22)

    r = total_row + 2
    for text, bold, color in [
        ("สองอย่างที่ต้องตรวจทุกครั้งก่อนเชื่อตารางสรุป", True, "9C5700"),
        ("หนึ่ง คอลัมน์ผลรวมกับคอลัมน์ผลรวมยืนยันด้วยวิธีที่สองต้องได้ค่าเท่ากันทุกแถว "
         "คอลัมน์แรกใช้ SUMIFS ส่วนคอลัมน์ที่สองใช้ SUMPRODUCT ซึ่งคำนวณคนละวิธีกันโดยสิ้นเชิง "
         "ถ้าสองค่านี้ต่างกัน แปลว่ามีข้อความปนอยู่ในคอลัมน์ตัวเลข หรือมีเซลล์ที่ยังไม่ได้ล้าง", False, "333333"),
        ("สอง ผลรวมของทุกกลุ่มต้องเท่ากับผลรวมทั้งหมดของคอลัมน์นั้นในชีตข้อมูลสะอาด "
         "และช่องร้อยละต้องรวมกันได้ 100 เปอร์เซ็นต์พอดี ถ้าได้น้อยกว่านั้นคือมีกลุ่มตกหล่น "
         "ซึ่งมักเกิดจากชื่อที่สะกดต่างกันเพียงช่องว่างเดียว", False, "333333"),
        ("", False, "333333"),
        ("แถวสีส้มที่ชื่อว่า ไม่ได้ระบุ มีความหมายอย่างไร", True, "9C5700"),
        ("คือรายการที่เซลล์คอลัมน์หมวดว่างเปล่า จึงไม่เข้ากลุ่มใดเลย ถ้าไม่แสดงแถวนี้ไว้ "
         "ผลรวมของทุกกลุ่มจะน้อยกว่าผลรวมทั้งหมดโดยไม่มีคำอธิบาย และร้อยละจะรวมกันไม่ถึง 100 "
         "PivotTable จริงจะแสดงแถวนี้ให้เองโดยใช้คำว่า ว่าง ซึ่งหลายคนเผลอลบทิ้งเพราะคิดว่าไม่สำคัญ", False, "333333"),
        ("", False, "333333"),
        ("เหตุผลที่ต้องมีคอลัมน์จำนวนรายการอยู่ข้างค่าเฉลี่ยเสมอ", True, "9C5700"),
        ("ค่าเฉลี่ยที่มาจากสองสามรายการแกว่งได้มาก การเห็นจำนวนรายการควบคู่กันทำให้ไม่สรุปเกินหลักฐาน "
         "นี่คือเกณฑ์ผ่านข้อหนึ่งของแบบฝึกหัด 2.3", False, "333333"),
    ]:
        ws.cell(row=r, column=1, value=text).font = Font(name=FONT, size=10, bold=bold, color=color)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        if text:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30 if len(text) > 80 else 18
        r += 1

    # ---------- ชีตขั้นตอนทำ PivotTable จริง ----------
    ws2 = wb.create_sheet("วิธีทำ PivotTable")
    ws2.column_dimensions["A"].width = 7
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 52
    ws2["A1"] = "ขั้นตอนทำ PivotTable จริงให้ได้ผลตรงกับชีตตารางสรุปแบบ Pivot"
    ws2["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws2["A2"] = ("ทำแล้วเทียบตัวเลขกับชีตก่อนหน้า ถ้าตรงกันทุกช่องแปลว่าทำถูก "
                 "ถ้าไม่ตรง ให้กลับไปดูว่าลากฟิลด์ผิดพื้นที่ หรือยังไม่ได้ล้างข้อมูล")
    ws2["A2"].font = Font(name=FONT, size=10, color="555555")
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A2:C2")
    ws2.row_dimensions[2].height = 30

    for j, h in enumerate(["ขั้น", "ใน Excel", "ใน Google Sheets"], start=1):
        ws2.cell(row=4, column=j, value=h)
    style_head(ws2, 3, row=4, fill=HEAD)

    steps = [
        ("คลิกเซลล์ใดก็ได้ในชีตข้อมูลสะอาด แล้วกด Ctrl+T เพื่อแปลงเป็นตารางแบบมีโครงสร้างก่อน "
         "ตั้งชื่อตารางให้สื่อความ",
         "เลือกช่วงข้อมูลทั้งหมด แล้วใช้เมนู จัดรูปแบบ เลือก แปลงเป็นตาราง"),
        ("แท็บ แทรก เลือก PivotTable แล้วเลือกวางในแผ่นงานใหม่ ห้ามวางทับข้างข้อมูลดิบ",
         "เมนู แทรก เลือก ตาราง Pivot แล้วเลือก แผ่นงานใหม่"),
        ("ลาก " + cat + " ลงพื้นที่ แถว",
         "ในแผงด้านขวา กด เพิ่ม ที่หัวข้อ แถว แล้วเลือก " + cat),
        ("ลาก " + num + " ลงพื้นที่ ค่า แล้วเปลี่ยนวิธีสรุปเป็น ผลรวม",
         "กด เพิ่ม ที่หัวข้อ ค่า เลือก " + num + " แล้วตั้ง สรุปโดย เป็น SUM"),
        ("ลาก " + num + " ลงพื้นที่ ค่า อีกครั้ง แล้วเปลี่ยนวิธีสรุปเป็น จำนวนนับ "
         "ขั้นนี้คือขั้นที่คนลืมบ่อยที่สุด แต่จำเป็นเพื่อไม่ให้สรุปเกินหลักฐาน",
         "กด เพิ่ม ที่ ค่า อีกครั้ง เลือกฟิลด์เดิม แล้วตั้ง สรุปโดย เป็น COUNTA"),
        ("ลาก " + num + " ลงพื้นที่ ค่า ครั้งที่สาม เปลี่ยนเป็น ค่าเฉลี่ย",
         "ทำซ้ำแล้วตั้ง สรุปโดย เป็น AVERAGE"),
        ("คลิกขวาที่ค่าผลรวม เลือก แสดงค่าเป็น แล้วเลือก ร้อยละของผลรวมทั้งหมด เพื่อได้คอลัมน์สุดท้าย",
         "ที่ฟิลด์ค่า เลือก แสดงเป็น แล้วเลือก ร้อยละของผลรวมทั้งหมด"),
        ("เพิ่ม Slicer จากแท็บ วิเคราะห์ PivotTable เพื่อให้คนอื่นกรองดูเองได้",
         "ใช้ แทรก เลือก ตัวควบคุมตัวกรอง แล้วเลือกคอลัมน์ที่ต้องการให้กรอง"),
        ("ทุกครั้งที่แก้ข้อมูลดิบ ต้องคลิกขวาที่ Pivot แล้วกด รีเฟรช ไม่อัปเดตเอง",
         "Sheets อัปเดตให้เองเมื่อข้อมูลต้นทางเปลี่ยน ข้อนี้ Sheets สะดวกกว่า Excel ชัดเจน"),
    ]
    r = 5
    for i, (ex, gs) in enumerate(steps, start=1):
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=ex)
        ws2.cell(row=r, column=3, value=gs)
        for j in range(1, 4):
            c = ws2.cell(row=r, column=j)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 42
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="เมื่อไรควรใช้ PivotTable และเมื่อไรควรใช้สูตร").font = Font(
        name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for text in [
        "ใช้ PivotTable เมื่อต้องการสำรวจข้อมูลและลองเปลี่ยนมุมมองเร็ว ๆ เพราะลากฟิลด์ทีเดียวเห็นผลทันที",
        "ใช้สูตรอย่าง SUMIFS เมื่อต้องการนำตัวเลขนั้นไปใช้ในสูตรอื่นต่อ หรือเมื่อต้องการให้ตัวเลขอัปเดตเองโดยไม่ต้องรีเฟรช",
        "ใช้ทั้งสองอย่างคู่กันเมื่อต้องตรวจงาน โดยให้สูตรเป็นตัวตรวจว่า PivotTable ที่ทำไว้ถูกต้อง ซึ่งเป็นวิธีที่ชีตนี้ใช้",
    ]:
        ws2.cell(row=r, column=1, value=text).font = Font(name=FONT, size=10, color="333333")
        ws2.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws2.row_dimensions[r].height = 28
        r += 1

    return {"หมวด": cat, "ตัวเลข": num, "กลุ่ม": len(groups), "แถวอ้างอิง": last}


# ============================================================ การจัดหน้าไฟล์เฉลย
def style_head(ws, ncol, row=1, fill=HEAD):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A{}".format(row + 1)


def autowidth(ws, maxw=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(w + 3, 10), maxw)


def title_block(ws, lines):
    r = 1
    for text, size, bold, color in lines:
        ws.cell(row=r, column=1, value=text).font = Font(
            name=FONT, size=size, bold=bold, color=color)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 20 if size < 14 else 26
        r += 1
    return r


# ============================================================ ประกอบไฟล์
def build(code):
    m = meta[code]
    src = openpyxl.load_workbook(os.path.join(SRC, m["fname"]), data_only=True)
    raw = src["ข้อมูลดิบ"]
    hdr = [c.value for c in raw[1]]
    rows = [list(r) for r in raw.iter_rows(min_row=2, values_only=True)]
    hdr = list(hdr)
    clean, stats = clean_sheet(hdr, rows)

    wb = openpyxl.Workbook()

    # ---------- ชีตแรก บอกว่านี่คือฉบับเฉลย ----------
    ws = wb.active
    ws.title = "อ่านก่อนเริ่ม"
    ws.column_dimensions["A"].width = 105
    r = title_block(ws, [
        ("ฉบับเฉลยสำหรับผู้สอน · " + m["title"], 14, True, "C00000"),
        ("ห้ามแจกไฟล์นี้ให้ผู้เรียนก่อนทำแบบฝึกหัดเสร็จ เพราะมีข้อมูลที่ทำความสะอาดแล้วและตัวเลขเฉลยครบ",
         11, True, "C00000"),
        ("", 11, False, "000000"),
        ("สาขา " + m["prog"], 11, True, NAVY),
        (m["context"], 10, False, "333333"),
        ("", 11, False, "000000"),
        ("ไฟล์นี้ประกอบด้วย", 11, True, NAVY),
        ("ชีต ข้อมูลสะอาด คือผลลัพธ์ที่ควรได้หลังทำความสะอาดครบทุกขั้น ใช้เทียบกับงานของผู้เรียนได้ทันที",
         10, False, "333333"),
        ("ชีต บันทึกการตัดสินใจ คือรายการว่าทำอะไรไปบ้าง กี่จุด และเพราะอะไร ใช้เป็นแม่แบบให้ผู้เรียนดู",
         10, False, "333333"),
        ("ชีต ตัวเลขเฉลย คือค่าที่ควรได้ตรงกัน ใช้ตรวจงานได้โดยไม่ต้องคำนวณเอง", 10, False, "333333"),
        ("", 11, False, "000000"),
        ("ข้อควรระวังในการตรวจงาน", 11, True, "9C5700"),
        ("ผู้เรียนอาจได้ตัวเลขต่างจากเฉลยเล็กน้อยโดยไม่ผิด ถ้าเลือกจัดการค่าที่หายไปด้วยวิธีอื่นและเขียนเหตุผลกำกับไว้ "
         "สิ่งที่ต้องตรวจคือมีบันทึกการตัดสินใจหรือไม่ ไม่ใช่ตัวเลขตรงกันหรือไม่", 10, False, "333333"),
    ])

    r += 1
    ws.cell(row=r, column=1, value="โจทย์ของสาขานี้แยกตามระดับ").font = Font(
        name=FONT, size=11, bold=True, color=NAVY)
    r += 1
    for lv, task, hint in m["tasks"]:
        ws.cell(row=r, column=1, value="[{}] {}".format(lv, task)).font = Font(
            name=FONT, size=10, bold=True, color="333333")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32
        r += 1
        ws.cell(row=r, column=1, value="สิ่งที่ผู้สอนควรเห็นในงาน: " + hint).font = Font(
            name=FONT, size=10, color="555555")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32
        r += 2

    # ---------- ข้อมูลสะอาด ----------
    ws = wb.create_sheet("ข้อมูลสะอาด")
    ws.append(hdr)
    for row in clean:
        ws.append(row)
    style_head(ws, len(hdr), fill=GREEN)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            if isinstance(c.value, (datetime.date, datetime.datetime)):
                c.number_format = "yyyy-mm-dd"
    autowidth(ws)

    # ---------- บันทึกการตัดสินใจ ----------
    ws = wb.create_sheet("บันทึกการตัดสินใจ")
    ws.append(["สิ่งที่ทำ", "จำนวนจุดที่แก้", "เหตุผล"])
    WHY = {
        "ตัดช่องว่างหน้าหลัง": "ช่องว่างที่มองไม่เห็นทำให้กลุ่มเดียวกันถูกนับเป็นคนละกลุ่มใน PivotTable และทำให้ฟังก์ชันค้นหาหาไม่พบ",
        "ตัดจุดหรืออักขระส่วนเกินท้ายข้อความ": "จุดท้ายชื่อเกิดจากนิสัยการพิมพ์ของผู้บันทึกแต่ละคน ไม่ใช่ส่วนหนึ่งของชื่อจริง",
        "แปลงตัวเลขที่เก็บเป็นข้อความ": "ตัวเลขที่เป็นข้อความจะไม่ถูกนำไปคำนวณ ทำให้ผลรวมและค่าเฉลี่ยต่ำกว่าความจริงโดยไม่มีการแจ้งเตือน",
        "แปลงตัวเลขที่มีเครื่องหมายคั่นหลักพัน": "เครื่องหมายคั่นทำให้โปรแกรมตีความเป็นข้อความ ต้องตัดออกก่อนจึงจะคำนวณได้",
        "แปลงวันที่ พ.ศ. แบบข้อความเป็นวันที่จริง": "วันที่ที่เป็นข้อความจะเรียงลำดับผิดและจัดกลุ่มตามเดือนไม่ได้",
        "ค่าแทนการไม่มีข้อมูลถูกเปลี่ยนเป็นช่องว่าง": "ค่าอย่าง -999 เป็นรหัสแทนช่วงที่ไม่มีข้อมูล ไม่ใช่ค่าที่วัดได้จริง ถ้าปล่อยไว้จะดึงค่าเฉลี่ยลง และถ้าแทนด้วยศูนย์จะยังผิดเหมือนเดิมแต่มองไม่ออก",
        "แยกหน่วยออกจากเซลล์ เหลือเฉพาะตัวเลข": "หน่วยควรอยู่ในชื่อคอลัมน์ ไม่ใช่ในเซลล์ เพื่อให้คำนวณได้โดยไม่เสียข้อมูลว่าหน่วยคืออะไร",
        "ลบแถวที่ซ้ำกันทั้งแถว": "แถวซ้ำเกิดจากการนำเข้าข้อมูลซ้ำรอบ ทำให้ผลรวมเกินจริงและจำนวนตัวอย่างผิด",
        "ลบแถวว่างทั้งแถว": "แถวว่างทำให้ช่วงข้อมูลขาดตอน ตัวกรองและ PivotTable จึงอ่านได้ไม่ครบ",
        "แปลงวันที่แบบข้อความเป็นวันที่จริง": "วันที่ที่เป็นข้อความจะเรียงลำดับผิดและจัดกลุ่มตามเดือนไม่ได้ แม้จะเขียนในรูปแบบสากลแล้วก็ตาม",
        "คงรหัสไว้เป็นข้อความโดยตั้งใจ ไม่แปลงเป็นตัวเลข": "นี่คือข้อยกเว้นสำคัญของกฎแปลงข้อความเป็นตัวเลข รหัสบัญชี รหัสไปรษณีย์ และเลขที่เอกสารเป็นตัวระบุ ไม่ใช่ปริมาณ ถ้าแปลงเป็นตัวเลขจะเสียเลขศูนย์นำหน้าและเชื่อมกับตารางอ้างอิงที่เก็บเป็นข้อความไม่ได้",
        "คงตัวเลขที่มีศูนย์นำหน้าไว้เป็นข้อความ": "เลขศูนย์นำหน้าจะหายไปทันทีที่แปลงเป็นตัวเลข ซึ่งแปลว่าข้อมูลเดิมสูญหายอย่างกู้คืนไม่ได้",
        "แปลงจำนวนในวงเล็บแบบงานบัญชีเป็นตัวเลขติดลบ": "ธรรมเนียมงานบัญชีเขียนจำนวนลบไว้ในวงเล็บ เช่น (1,250.00) หมายถึงลบหนึ่งพันสองร้อยห้าสิบ ไม่ใช่ข้อความ ถ้าปล่อยเป็นข้อความจะคำนวณไม่ได้ และถ้าตัดวงเล็บออกโดยไม่ใส่เครื่องหมายลบ ยอดจะกลับข้างทั้งคอลัมน์",
    }
    total = 0
    for k, n in sorted(stats.items(), key=lambda x: -x[1]):
        why = WHY.get(k)
        if why is None and k.startswith("รวมชื่อกลุ่มที่ต่างกันเพียงตัวพิมพ์"):
            why = ("SUMIFS และ COUNTIFS ของ Excel ไม่แยกตัวพิมพ์ใหญ่เล็ก จึงนับ Pm-ls กับ Pm-Ls เป็นกลุ่มเดียวกัน "
                   "แต่ถ้าปล่อยให้มีสองสะกดอยู่ในตารางสรุป ข้อมูลชุดเดียวกันจะถูกนับสองครั้ง "
                   "และผลรวมของทุกกลุ่มจะเกินผลรวมทั้งหมด ตรวจพบได้จากช่องร้อยละที่เกิน 100 เปอร์เซ็นต์")
        elif why is None and k.startswith("รวมชื่อกลุ่ม"):
            why = "ชื่อที่ต่างกันเพียงช่องว่างหรือจุด คือชื่อเดียวกันในความหมาย จึงรวมเป็นรูปแบบที่พบบ่อยที่สุด"
        elif why is None and k.startswith("ย้ายข้อความที่แปลงเป็นตัวเลขไม่ได้"):
            why = ("ข้อความอย่าง หลายตัว หรือ 2-3 หรือ ไม่แน่ชัด ไม่มีค่าตัวเลขที่แน่นอน "
                   "การเดาแทนผู้บันทึกภาคสนามเป็นการสร้างข้อมูลขึ้นมาเอง จึงย้ายไปเก็บในคอลัมน์หมายเหตุ "
                   "แล้วปล่อยช่องตัวเลขให้ว่าง ข้อมูลเดิมไม่สูญหายและคอลัมน์ตัวเลขคำนวณได้ทั้งคอลัมน์ "
                   "ในรายงานต้องระบุว่ามีกี่รายการที่ไม่ได้นำมาคำนวณและเพราะอะไร")
        ws.append([k, n, why or "ปรับให้เป็นรูปแบบมาตรฐานเดียวกันทั้งไฟล์"])
        total += n
    ws.append(["รวมทุกจุดที่แก้", total, "จำนวนแถวคงเหลือหลังทำความสะอาด {} แถว จากเดิม {} แถว".format(
        len(clean), len(rows))])
    style_head(ws, 3, fill=ORANGE)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 74
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)

    # ---------- ตัวเลขเฉลย ----------
    ws = wb.create_sheet("ตัวเลขเฉลย")
    ws.append(["รายการ", "ค่า"])
    def flat(v):
        if isinstance(v, dict):
            return " · ".join("{} = {}".format(a, b) for a, b in v.items())
        if isinstance(v, (list, tuple)):
            return ", ".join(map(str, v)) if v else "ไม่มี"
        return v
    for k, v in KEY[code].items():
        ws.append([k, flat(v)])
    style_head(ws, 2, fill=HEAD)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
    # สถิติที่คำนวณจากชีตข้อมูลสะอาดโดยตรง ใช้ตรวจงานผู้เรียนได้ทันที
    ws.append([])
    ws.append(["สถิติของชีตข้อมูลสะอาด ใช้เทียบกับงานผู้เรียนได้ทันที", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.append(["คอลัมน์ตัวเลข", "จำนวนค่าที่มี · ค่าเฉลี่ย · ค่าต่ำสุด · ค่าสูงสุด"])
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.cell(row=ws.max_row, column=2).font = Font(name=FONT, size=10, bold=True, color=NAVY)
    for j, name in enumerate(hdr):
        vals = [r[j] for r in clean if isinstance(r[j], (int, float))]
        if len(vals) < len(clean) * 0.5 or not vals:
            continue
        ws.append([name, "{:,} ค่า · เฉลี่ย {:,.2f} · ต่ำสุด {:,.2f} · สูงสุด {:,.2f}".format(
            len(vals), sum(vals) / len(vals), min(vals), max(vals))])
    ws.append(["หมายเหตุสำคัญ",
               "ตัวเลขในตารางบนคำนวณจากข้อมูลดิบเพื่อแสดงผลของการไม่ล้างข้อมูล "
               "ส่วนตารางล่างคำนวณจากข้อมูลที่ล้างครบทุกขั้นแล้วซึ่งลบแถวซ้ำออกด้วย "
               "ค่าเฉลี่ยสองชุดจึงต่างกันเล็กน้อยได้ตามปกติ ให้ใช้ตารางล่างเป็นเกณฑ์ตรวจงาน"])
    ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, size=10, bold=True, color="9C5700")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.font.size is None or not c.font.bold:
                c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 78

    # ---------- ตารางสรุปแบบ Pivot และวิธีทำ ----------
    pv = add_pivot_sheets(wb, code, hdr, clean)

    # ---------- ตารางอ้างอิงคัดลอกมาทั้งชุด ----------
    for ref in m["refs"]:
        s0 = src[ref["sheet"]]
        ws = wb.create_sheet(ref["sheet"])
        for row in s0.iter_rows(values_only=True):
            ws.append(list(row))
        style_head(ws, s0.max_column, fill=GREY)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name=FONT, size=10)
                c.border = BORDER
        autowidth(ws)

    fn = m["fname"].replace(".xlsx", "_เฉลย.xlsx")
    wb.save(os.path.join(OUT, fn))

    # ---------- เวอร์ชัน .csv สำหรับนำเข้า Google Sheets ----------
    import csv
    for label, sheet_hdr, sheet_rows in [("ข้อมูลดิบ", hdr, rows), ("เฉลยข้อมูลสะอาด", hdr, clean)]:
        base = m["fname"].replace(".xlsx", "")
        cn = "{}_{}.csv".format(base, label)
        with open(os.path.join(CSVDIR, cn), "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(sheet_hdr)
            for row in sheet_rows:
                w.writerow(["" if v is None else
                            (v.strftime("%Y-%m-%d") if isinstance(v, (datetime.date, datetime.datetime)) else v)
                            for v in row])
    for ref in m["refs"]:
        s0 = src[ref["sheet"]]
        cn = "{}_{}.csv".format(m["fname"].replace(".xlsx", ""), ref["sheet"])
        with open(os.path.join(CSVDIR, cn), "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            for row in s0.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])

    return fn, len(rows), len(clean), total, dict(stats), pv


report = {}
for code in ["AG", "AC", "BA", "CB", "ED", "FT", "GS"]:
    fn, n0, n1, total, st, pv = build(code)
    report[code] = {"ไฟล์": fn, "แถวก่อน": n0, "แถวหลัง": n1, "จุดที่แก้": total,
                "รายละเอียด": st, "ตารางสรุป": pv}
    print("{}  {:>4} -> {:>4} แถว  แก้ {:>4} จุด  สรุปตาม {} {} กลุ่ม".format(
        code, n0, n1, total, pv["หมวด"], pv["กลุ่ม"]))
json.dump(report, open("answers_report.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
