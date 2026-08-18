#!/usr/bin/env python3
"""ไฟล์รายชื่อผู้รับสำหรับจดหมายเวียน ตั้งใจใส่ปัญหาที่พบจริงไว้ให้ผู้เรียนเจอ
เพราะจดหมายเวียนที่พังส่วนใหญ่ไม่ได้พังที่ Word แต่พังที่ไฟล์รายชื่อ"""
import sys, random, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1]
FONT = "Tahoma"
NAVY = "1F3864"; HEAD = "D9E2F3"; ORANGE = "FBE5D6"; WARN = "FCEAE8"
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
random.seed(2569)

COLS = ["คำนำหน้า", "ชื่อ", "นามสกุล", "ตำแหน่ง", "หน่วยงาน", "อีเมล",
        "ระดับที่อบรม", "วันที่อบรม", "จำนวนชั่วโมง",
        "คะแนนก่อนเรียน", "คะแนนหลังเรียน", "ผลการประเมิน",
        "เลขที่หนังสือ", "เรื่อง", "วันที่หนังสือ"]

FIRST = ["สมชาย", "สุภาพร", "ธนกฤต", "ปิยะดา", "อนุชา", "กมลชนก", "ณัฐพงษ์", "วรรณิดา",
         "ศิริชัย", "พัชราภา", "ธีรเดช", "อารยา", "กิตติศักดิ์", "เบญจวรรณ", "รัฐพล",
         "นภัสสร", "จิรวัฒน์", "ชนิดา"]
LAST = ["ใจดี", "แสงทอง", "บุญมี", "รักไทย", "ศรีสุข", "วงศ์คำ", "ทองอินทร์", "พูนสิน",
        "เกษมสุข", "อินทรีย์", "ชัยมงคล", "ธนบดี", "สุวรรณ", "พรหมมา", "ยิ่งเจริญ",
        "มั่งมี", "เรืองศรี", "กล้าหาญ"]
POS = ["นักวิชาการเกษตร", "อาจารย์", "นักวิทยาศาสตร์", "เจ้าหน้าที่วิเคราะห์นโยบายและแผน",
       "นักวิชาการเงินและบัญชี", "นักวิชาการศึกษา"]
ORG = ["สาขาวิชาวิทยาศาสตร์การเกษตร", "สาขาวิชาเทคโนโลยีการอาหาร", "งานคลังและพัสดุ",
       "สาขาวิชาชีววิทยาเชิงอนุรักษ์", "งานบริการการศึกษา", "สาขาวิชาธรณีศาสตร์"]
LEVEL = ["Beginner", "Intermediate", "Advanced"]

wb = Workbook()

# ---------------- ชีตคำอธิบาย ----------------
ws = wb.active
ws.title = "อ่านก่อนเริ่ม"
ws.column_dimensions["A"].width = 108
lines = [
    ("ไฟล์รายชื่อผู้รับสำหรับฝึกทำจดหมายเวียน", 14, True, NAVY),
    ("ใช้คู่กับไฟล์ แม่แบบจดหมายเวียน_หนังสือรับรองการอบรม.docx", 11, False, "333333"),
    ("", 11, False, "000000"),
    ("โครงสร้างที่จดหมายเวียนต้องการ", 12, True, NAVY),
    ("ชีต รายชื่อ มีหัวตารางแถวเดียวอยู่แถวที่ 1 และข้อมูลเริ่มแถวที่ 2 ต่อเนื่องไม่มีแถวว่างคั่น", 10, False, "333333"),
    ("ถ้าหัวตารางไม่อยู่แถวแรก Word จะอ่านชื่อคอลัมน์ผิดทั้งหมด และรายการเขตข้อมูลจะกลายเป็น F1 F2 F3", 10, False, "333333"),
    ("ชื่อคอลัมน์ต้องไม่มีช่องว่างนำหน้าหรือตามหลัง เพราะจะทำให้จับคู่เขตข้อมูลไม่ได้", 10, False, "333333"),
    ("", 11, False, "000000"),
    ("ไฟล์นี้มีปัญหาที่ตั้งใจใส่ไว้ห้าอย่าง ให้หาให้ครบก่อนเริ่มผสาน", 12, True, "C00000"),
    ("ทั้งห้าอย่างเป็นปัญหาที่เจอจริงเมื่อรับไฟล์รายชื่อมาจากผู้อื่น ไม่ใช่ปัญหาที่สร้างมาเพื่อแกล้ง", 10, False, "333333"),
    ("ถ้าเริ่มผสานก่อนแก้ จดหมายจะออกมาผิดโดยที่ Word ไม่แจ้งเตือนอะไรเลย", 10, False, "C00000"),
    ("", 11, False, "000000"),
    ("สิ่งที่ต้องส่งเป็นชิ้นงาน", 12, True, NAVY),
    ("หนึ่ง ไฟล์รายชื่อที่แก้แล้ว สอง ไฟล์จดหมายที่ผสานครบทุกฉบับในไฟล์เดียว", 10, False, "333333"),
    ("สาม บันทึกสั้น ๆ ว่าพบปัญหาอะไรบ้างและแก้อย่างไร โดยเฉพาะข้อที่ตัดสินใจไม่ส่งจดหมายให้บางคน", 10, False, "333333"),
]
r = 1
for t, sz, bold, col in lines:
    c = ws.cell(row=r, column=1, value=t)
    c.font = Font(name=FONT, size=sz, bold=bold, color=col)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 22 if sz < 13 else 28
    r += 1

# ---------------- ชีตรายชื่อ ----------------
ws = wb.create_sheet("รายชื่อ")
rows = []
for i in range(18):
    lv = LEVEL[i % 3]
    hours = {"Beginner": 3, "Intermediate": 3, "Advanced": 3}[lv]
    pre = random.randint(28, 62)
    post = pre + random.randint(12, 34)
    post = min(post, 98)
    grade = "ดีมาก" if post >= 85 else ("ดี" if post >= 70 else "ผ่าน")
    day = 8 + (i % 3) * 7
    rows.append([
        random.choice(["นาย", "นาง", "นางสาว"]),
        FIRST[i], LAST[i], random.choice(POS), random.choice(ORG),
        "user{:02d}@mahidol.ac.th".format(i + 1),
        lv, datetime.date(2026, 9, day), hours,
        pre, round(post + random.random(), 6), grade,
        "{:04d}".format(101 + i), "ขอส่งหนังสือรับรองการเข้าร่วมอบรม", datetime.date(2026, 9, 30)
    ])

# ---------- กับดักที่ตั้งใจใส่ ----------
TRAPS = []
# 1 ช่องว่างต่อท้ายชื่อ ทำให้จดหมายมีช่องว่างเกินหน้านามสกุล
rows[2][1] = rows[2][1] + " "
rows[7][2] = " " + rows[7][2]
TRAPS.append(("ช่องว่างนำหน้าหรือตามหลังชื่อ", 2, "จดหมายจะมีช่องว่างเกินหรือขาด ดูไม่เป็นทางการ ใช้ TRIM แก้"))
# 2 อีเมลซ้ำ คนเดียวได้จดหมายสองฉบับ
rows[11][5] = rows[4][5]
TRAPS.append(("อีเมลซ้ำกันสองแถว", 1, "คนเดียวจะได้จดหมายสองฉบับ ต้องตรวจซ้ำด้วย COUNTIF ก่อนผสาน"))
# 3 คำนำหน้าว่าง ทำให้ประโยคขาด
rows[5][0] = None
TRAPS.append(("คำนำหน้าว่าง", 1, "ประโยคในจดหมายจะขาดคำ ต้องเติมให้ครบหรือใช้ IF ในเขตข้อมูลจัดการ"))
# 4 วันที่เป็นข้อความ พ.ศ.
rows[9][7] = "22/09/2569"
rows[14][7] = "15/09/2569"
TRAPS.append(("วันที่เป็นข้อความแบบ พ.ศ.", 2, "Word จะพิมพ์ออกมาตามที่เห็นเป็นข้อความ และรหัสรูปแบบวันที่จะใช้ไม่ได้"))
# 5 คะแนนเป็นข้อความมีเครื่องหมายเปอร์เซ็นต์ปน
rows[3][10] = str(int(rows[3][10])) + "%"
rows[16][10] = str(int(rows[16][10])) + " เปอร์เซ็นต์"
TRAPS.append(("คะแนนมีหน่วยปนในเซลล์", 2, "คำนวณต่อไม่ได้ และรหัสรูปแบบตัวเลขจะไม่ทำงาน ต้องเก็บเฉพาะตัวเลข"))

ws.append(COLS)
for row in rows:
    ws.append(row)
for j in range(1, len(COLS) + 1):
    c = ws.cell(row=1, column=j)
    c.font = Font(name=FONT, bold=True, size=10, color=NAVY)
    c.fill = PatternFill("solid", fgColor=HEAD)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER
ws.freeze_panes = "A2"
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        if isinstance(c.value, datetime.date):
            c.number_format = "yyyy-mm-dd"
for col in ws.columns:
    letter = get_column_letter(col[0].column)
    w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
    ws.column_dimensions[letter].width = min(max(w + 3, 11), 34)

# ---------------- ชีตเฉลยกับดัก ----------------
ws = wb.create_sheet("เฉลยกับดัก (ผู้สอน)")
ws.append(["ปัญหาที่ซ่อนไว้", "จำนวนจุด", "ผลที่เกิดถ้าไม่แก้ และวิธีแก้"])
for t, n, why in TRAPS:
    ws.append([t, n, why])
ws.append(["รวม", sum(t[1] for t in TRAPS), "ห้าประเภท รวม {} จุด".format(sum(t[1] for t in TRAPS))])
for j in range(1, 4):
    c = ws.cell(row=1, column=j)
    c.font = Font(name=FONT, bold=True, size=10, color=NAVY)
    c.fill = PatternFill("solid", fgColor=ORANGE)
    c.border = BORDER
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=ws.max_row, column=1).font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 78
ws.freeze_panes = "A2"

wb.save(OUT)
print("เขียนแล้ว", OUT)
print("แถวรายชื่อ", len(rows), "· คอลัมน์", len(COLS), "· กับดัก", len(TRAPS), "ประเภท", sum(t[1] for t in TRAPS), "จุด")
